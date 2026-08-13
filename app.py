import os
import re
import sys
import json
import glob
import uuid
import queue
import threading
import subprocess
from flask import Flask, render_template, request, Response, jsonify, send_from_directory, abort
import time
from datetime import datetime

# In a frozen PyInstaller build, __file__ resolves inside the bundled _internal/ folder, not
# the folder actually holding app.exe (and the sibling worker .exes _worker_cmd looks for) —
# os.path.dirname(sys.executable) is the correct base for WRITABLE data (accounts.json,
# runs/, screenshots/, ...) in that case. Bundled READ-ONLY resources (templates/, static/)
# live under sys._MEIPASS instead (PyInstaller's _internal/ folder in --onedir builds).
# Dev mode (running app.py from source) is unaffected: sys.frozen is unset, both point at the
# same project root as today.
if getattr(sys, "frozen", False):
    BASE_DIR = os.path.dirname(sys.executable)
    RESOURCE_DIR = sys._MEIPASS
else:
    BASE_DIR = os.path.dirname(os.path.abspath(__file__))
    RESOURCE_DIR = BASE_DIR

app = Flask(__name__, root_path=BASE_DIR,
           template_folder=os.path.join(RESOURCE_DIR, "templates"),
           static_folder=os.path.join(RESOURCE_DIR, "static"))
SCREENSHOTS_DIR = os.path.join(BASE_DIR, "screenshots")
RECORDINGS_DIR = os.path.join(BASE_DIR, "recordings")
RUNS_DIR = os.path.join(BASE_DIR, "runs")
RUNS_INDEX = os.path.join(RUNS_DIR, "index.json")
ACCOUNTS_FILE = os.path.join(BASE_DIR, "accounts.json")
RESULTS_FILE = os.path.join(BASE_DIR, "test_results.json")
SCHEDULES_FILE = os.path.join(BASE_DIR, "scheduled_jobs.json")
SCHEDULED_UPLOADS_DIR = os.path.join(RUNS_DIR, "scheduled")

# Ensure dirs exist
os.makedirs(SCREENSHOTS_DIR, exist_ok=True)
os.makedirs(RECORDINGS_DIR, exist_ok=True)
os.makedirs(RUNS_DIR, exist_ok=True)
os.makedirs(SCHEDULED_UPLOADS_DIR, exist_ok=True)


def _slug(s, maxlen=40):
    """Filesystem-safe slug for run-folder names."""
    s = re.sub(r"[^A-Za-z0-9._-]+", "-", (s or "").strip()).strip("-")
    return (s[:maxlen] or "game")


def _worker_cmd(script, *args):
    """Build the argv for a worker subprocess (test_spin_button.py, crash_auto.py, or
    tlogs_validate.py). In dev mode (running from source) this shells out to the .py file
    via the SAME interpreter running app.py, exactly as before. In a frozen/compiled build
    (see build_exe.py) a frozen app.exe can't be told to run an arbitrary .py file the way
    `python -u script.py` can — so each worker is compiled into its OWN executable sitting
    next to app.exe, and this just points at that instead."""
    if getattr(sys, "frozen", False):
        exe = os.path.join(BASE_DIR, os.path.splitext(script)[0] + ".exe")
        return [exe, *args]
    return [sys.executable, "-u", script, *args]


def _parse_money(text):
    """Best-effort numeric parse of a currency-formatted amount (the crash catalog's
    minBetAmount comes back as whatever the provider set — '1', '0.50', 'R 1', 'R0.50' have
    all been observed live — not a bare number). Deliberately NOT test_spin_button.parse_amount:
    importing that module here would pull in its heavy import-time side effects (a Gemini
    client, Playwright) into the Flask process for a one-line strip-and-parse."""
    if text is None:
        return None
    digits = re.sub(r"[^\d.]", "", str(text))
    try:
        return float(digits) if digits else None
    except ValueError:
        return None

# Single-user global run state. A "run" is 1..N worker subprocesses (parallel DSC gives
# each selected account its own browser worker); their stdout lines are multiplexed into
# one queue that /stream drains, prefixed [W1]/[W2]/… when there's more than one worker.
WORKERS = []          # [{"proc": Popen, "label": "W1"}]
LOG_Q = None          # queue.Queue for the current run; None sentinel = stream end
CURRENT_RUN = {"game": None, "start_time": None, "status": "idle"}
# Single source of truth for the live log: every worker line is appended here and
# /stream serves each client from its own read position. That makes the stream
# broadcast-safe — any number of tabs get the FULL log (the old queue was
# consume-once: with two dashboard tabs open, each tab got random halves).
# Reset per run. RUN_DONE flips when the watcher has appended the final lines.
from collections import deque
LOG_HISTORY = deque(maxlen=20000)
RUN_DONE = {"v": True}


def _log_put(q, line):
    LOG_HISTORY.append(line)


def _workers_alive():
    return any(w["proc"].poll() is None for w in WORKERS)


def _stop_workers():
    for w in WORKERS:
        if w["proc"].poll() is None:
            w["proc"].terminate()
            w["proc"].wait()


def _pump_worker(proc, label, q, drop_payload):
    """Reader thread: one per worker. In batch mode the per-game REPORTPAYLOAD blocks
    are dropped — 300 inline JSON reports would swamp the log pane, and each game's
    results.json is already persisted in its run folder."""
    prefix = f"[{label}] " if label else ""
    in_payload = False
    for line in iter(proc.stdout.readline, ''):
        line = line.rstrip("\r\n")
        if drop_payload:
            s = line.strip()
            if s == "REPORTPAYLOAD===":
                in_payload = True
                continue
            if s == "===REPORTPAYLOAD":
                in_payload = False
                continue
            if in_payload:
                continue
        _log_put(q, prefix + line)
    proc.stdout.close()
    proc.wait()


def _finish_batch(q, batch):
    """Shared batch-completion tail (report summary + sweep DB recording), used once every
    worker has finished — by both the immediate-launch path and the throttled scheduler
    below, so the two launch styles report exactly the same way."""
    if batch and batch.get("report") and os.path.exists(batch["report"]):
        try:
            _log_put(q, "[BATCH] " + _report_summary(batch["report"]))
        except Exception as e:
            _log_put(q, f"[BATCH] summary unavailable: {e}")
        _log_put(q, f"[BATCH] Complete — report: {os.path.basename(batch['report'])}")
    # Auto-sweep: record each tested game into the prod DB so the weekly rotation holds
    # (unless the run was flagged do-not-record). Match picks to their result rows by name.
    if batch and batch.get("sweep"):
        _record_sweep(batch["sweep"], batch.get("report"), q)


def _watch_workers(workers, q, batch):
    """Waits for every worker, then closes the stream (and, for batches, appends a
    result summary read back from the shared report)."""
    for w in workers:
        w["proc"].wait()
    _finish_batch(q, batch)
    RUN_DONE["v"] = True
    q.put(None)


def _record_sweep(sweep, report_path, q):
    """Persist an auto-sweep's results to dsc_history.db (honours the do-not-record flag)."""
    try:
        from modules import dsc_history_db
        if not sweep.get("record"):
            _log_put(q, "[SWEEP] do-not-record: results NOT written to the prod database")
            return
        by_name = {}
        rec_path = (report_path or "").replace(".xlsx", "_records.jsonl")
        if rec_path and os.path.exists(rec_path):
            with open(rec_path, encoding="utf-8") as fh:
                for line in fh:
                    line = line.strip()
                    if line:
                        r = json.loads(line)
                        by_name[(r.get("game") or "").strip().lower()] = r
        n = 0
        for p in sweep.get("picks", []):
            out = by_name.get((p["name"] or "").strip().lower(), {})
            dsc_history_db.record_test(sweep["brand"], sweep["region"], p["provider"],
                                       p["id"], p["name"], out, run_id=sweep.get("run_id", ""),
                                       record=True)
            n += 1
        _log_put(q, f"[SWEEP] Recorded {n} result(s) to the prod database (dsc_history.db)")
    except Exception as e:
        _log_put(q, f"[SWEEP] DB record failed: {e}")


def _report_summary(report_path):
    """One-line Pass/Fail tally straight from the report sheet."""
    from openpyxl import load_workbook
    from modules import dsc_report as dr
    wb = load_workbook(report_path)
    ws, hdr, cols = dr._pick_sheet(wb)   # header may not be row 1 (title rows, dead tabs)

    def idx(canonical):
        c = dr._col(cols, canonical)
        return c - 1 if c else None

    name_i, launch_i = idx("Game Name"), idx("Launch")
    tally = {"Launch": [0, 0], "Bet Placed": [0, 0], "Tlogs": [0, 0]}
    pending = 0
    awaiting = 0   # Tlogs "Pending": spun, but transaction validation hasn't run yet
    for row in ws.iter_rows(min_row=hdr + 1, values_only=True):
        if name_i is None or name_i >= len(row) or not str(row[name_i] or "").strip():
            continue   # filler/blank rows aren't games
        if launch_i is None or launch_i >= len(row):
            continue
        if not str(row[launch_i] or "").strip():
            pending += 1
            continue
        for col, counts in tally.items():
            i = idx(col)
            v = str(row[i] or "").strip().lower() if i is not None and i < len(row) else ""
            if v == "pass":
                counts[0] += 1
            elif v == "fail":
                counts[1] += 1
            elif v == "pending" and col == "Tlogs":
                awaiting += 1
    wb.close()
    parts = [f"{col} {p} Pass / {f} Fail" for col, (p, f) in tally.items()]
    if awaiting:
        parts.append(f"{awaiting} awaiting Tlogs validation")
    if pending:
        parts.append(f"{pending} not run")
    return "Summary: " + " · ".join(parts)


def _start_workers(cmds, batch=None, max_parallel=None):
    """Spawn one subprocess per (label, cmd[, env]), wire reader threads into a fresh queue.
    max_parallel=None (default): launch every worker immediately — unchanged behavior, used by
    the single-game launch (always 1 worker) and anywhere a small, fixed worker count is fine.
    Pass max_parallel to cap how many run at once; the rest queue and start as slots free, the
    same wave-scheduling _start_fleet already uses for the auto-sweep — added so a manual batch
    launch with many worker accounts (e.g. 34) doesn't open that many real browsers at the same
    instant (CPU/RAM contention, plus every worker's Gemini vision calls colliding on the single
    shared API key at once)."""
    global WORKERS, LOG_Q
    _stop_workers()
    FLEET_STOP.clear()
    LOG_Q = queue.Queue()
    LOG_HISTORY.clear()
    RUN_DONE["v"] = False
    WORKERS = []
    drop_payload = batch is not None
    q = LOG_Q

    if not max_parallel or max_parallel >= len(cmds):
        for item in cmds:
            label, cmd = item[0], item[1]
            env = item[2] if len(item) > 2 else None
            proc = subprocess.Popen(
                cmd, stdout=subprocess.PIPE, stderr=subprocess.STDOUT,
                text=True, encoding="utf-8", errors="replace", bufsize=1, cwd=BASE_DIR,
                env=env
            )
            WORKERS.append({"proc": proc, "label": label})
            threading.Thread(target=_pump_worker, args=(proc, label, LOG_Q, drop_payload),
                             daemon=True).start()
        threading.Thread(target=_watch_workers, args=(list(WORKERS), LOG_Q, batch),
                         daemon=True).start()
        return

    def scheduler():
        pending = list(cmds)
        running = []   # {"proc", "thread", "finished_at": float|None}
        while (pending or running) and not FLEET_STOP.is_set():
            occupied = sum(1 for r in running if r["finished_at"] is None)
            while pending and occupied < max_parallel and not FLEET_STOP.is_set():
                item = pending.pop(0)
                label, cmd = item[0], item[1]
                env = item[2] if len(item) > 2 else None
                try:
                    proc = subprocess.Popen(
                        cmd, stdout=subprocess.PIPE, stderr=subprocess.STDOUT,
                        text=True, encoding="utf-8", errors="replace", bufsize=1,
                        cwd=BASE_DIR, env=env)
                except Exception as e:
                    _log_put(q, f"[BATCH] worker {label} failed to start: {e}")
                    continue
                WORKERS.append({"proc": proc, "label": label})
                th = threading.Thread(target=_pump_worker, args=(proc, label, q, drop_payload),
                                      daemon=True)
                th.start()
                running.append({"proc": proc, "thread": th, "finished_at": None, "label": label})
                occupied += 1
            time.sleep(0.5)
            for r in running:
                if r["finished_at"] is None and r["proc"].poll() is not None:
                    r["thread"].join(timeout=5)
                    r["finished_at"] = time.time()
            running = [r for r in running if r["finished_at"] is None]
        _finish_batch(q, batch)
        RUN_DONE["v"] = True
        q.put(None)

    threading.Thread(target=scheduler, daemon=True).start()


FLEET_STOP = threading.Event()   # set by /stop so a capped fleet stops launching queued waves


def _start_fleet(worker_specs, max_parallel, group_meta, cooldown=0):
    """Run a multi-region auto-sweep as a fleet of browser workers, at most `max_parallel`
    running at once (queued specs start as slots free — so 'all products' across 8 regions
    doesn't open 8 browsers at once). `worker_specs` = [{label, cmd, env}]; `group_meta` =
    one entry per (brand, region) with its shared report + picks, recorded to the prod DB when
    ALL workers finish (each honours its own do-not-record flag via _record_sweep).

    `cooldown` (seconds, default 0 — unused by the slot Auto Sweep, which spreads across
    DIFFERENT accounts and has no reason to wait): holds a finished worker's slot occupied for
    this long before the next pending spec may start in it. Added for the crash sweep
    (max_parallel=1, SAME account every launch) after the 2026-08-03 provider validation sweep
    showed frequent 'session held by another tab' / DISCONNECTED aborts on games launched
    immediately after a previous game's browser closed — our own launch-URL fetch is always
    fresh (no client-side token reuse), so this points at the casino backend's own wallet/game
    session lock for that account not having released yet, especially after a run that aborted
    without a clean in-game exit. A gap before requesting the next launch gives that lock time
    to expire instead of guaranteed-colliding with it."""
    global WORKERS, LOG_Q
    _stop_workers()
    FLEET_STOP.clear()
    LOG_Q = queue.Queue()
    LOG_HISTORY.clear()
    RUN_DONE["v"] = False
    WORKERS = []
    q = LOG_Q

    def scheduler():
        pending = list(worker_specs)
        running = []   # {"proc", "thread", "finished_at": float|None}
        while (pending or running) and not FLEET_STOP.is_set():
            occupied = sum(1 for r in running
                          if r["finished_at"] is None or time.time() - r["finished_at"] < cooldown)
            while pending and occupied < max_parallel and not FLEET_STOP.is_set():
                spec = pending.pop(0)
                try:
                    proc = subprocess.Popen(
                        spec["cmd"], stdout=subprocess.PIPE, stderr=subprocess.STDOUT,
                        text=True, encoding="utf-8", errors="replace", bufsize=1,
                        cwd=BASE_DIR, env=spec.get("env"))
                except Exception as e:
                    _log_put(q, f"[SWEEP] worker {spec['label']} failed to start: {e}")
                    continue
                w = {"proc": proc, "label": spec["label"]}
                WORKERS.append(w)
                th = threading.Thread(target=_pump_worker, args=(proc, spec["label"], q, True),
                                      daemon=True)
                th.start()
                running.append({"proc": proc, "thread": th, "finished_at": None,
                                "label": spec["label"]})
                occupied += 1
            time.sleep(0.5)
            for r in running:
                if r["finished_at"] is None and r["proc"].poll() is not None:
                    r["thread"].join(timeout=5)
                    r["finished_at"] = time.time()
                    if cooldown:
                        _log_put(q, f"[SWEEP] {r['label']} finished — "
                                    f"{cooldown:g}s cooldown before the next launch...")
            running = [r for r in running
                      if r["finished_at"] is None or time.time() - r["finished_at"] < cooldown]
        # Every worker finished (or the run was stopped): tally + record each region.
        for gm in group_meta:
            try:
                if gm.get("report") and os.path.exists(gm["report"]):
                    _log_put(q, f"[SWEEP] {gm['brand']} {gm['region']}: "
                                + _report_summary(gm["report"]))
            except Exception as e:
                _log_put(q, f"[SWEEP] {gm['brand']} {gm['region']} summary unavailable: {e}")
            _record_sweep(gm, gm.get("report"), q)
        RUN_DONE["v"] = True
        q.put(None)

    threading.Thread(target=scheduler, daemon=True).start()

# Cache of auth tokens so the typeahead doesn't re-authenticate on every keystroke.
# { username: (token, epoch_seconds) }; entries expire after TOKEN_TTL.
_TOKEN_CACHE = {}
TOKEN_TTL = 600  # 10 minutes


def _get_token(username, password, brand="betway", region="ZA"):
    """Return a cached token for the (account, brand, region), authenticating if needed."""
    key = (username, brand, region)
    hit = _TOKEN_CACHE.get(key)
    if hit and (time.time() - hit[1]) < TOKEN_TTL:
        return hit[0], None
    from modules.auth_handler import AuthHandler
    auth = AuthHandler().authenticate(username, password, brand=brand, region=region)
    if auth.get("success"):
        _TOKEN_CACHE[key] = (auth["token"], time.time())
        return auth["token"], None
    return None, auth.get("message", "auth failed")


def load_accounts():
    if os.path.exists(ACCOUNTS_FILE):
        with open(ACCOUNTS_FILE, "r") as f:
            return json.load(f)
    return []


def save_accounts(accounts):
    with open(ACCOUNTS_FILE, "w") as f:
        json.dump(accounts, f, indent=2)


# ─── Scheduled jobs ──────────────────────────────────────────────
# Same flat-JSON-list pattern as accounts.json/load_accounts/save_accounts — always read fresh
# from disk (no in-memory cache), which is what gives this restart-safety for free: nothing
# special has to happen at startup, the next poll tick just re-reads whatever's on disk.
def load_schedules():
    if os.path.exists(SCHEDULES_FILE):
        with open(SCHEDULES_FILE, "r", encoding="utf-8") as f:
            return json.load(f)
    return []


def save_schedules(schedules):
    with open(SCHEDULES_FILE, "w", encoding="utf-8") as f:
        json.dump(schedules, f, indent=2)


def _schedule_is_due(job, now):
    """Pure due-check (no side effects) so the poll loop's decision is easy to reason about.
    "once": fires exactly once, at/after run_at. "daily": fires once per matching day, at/after
    time_of_day (days_of_week empty/absent = every day)."""
    rec = job.get("recurrence") or {}
    kind = rec.get("kind")
    last_run_at = job.get("last_run_at")
    if kind == "once":
        run_at = rec.get("run_at")
        if not run_at or last_run_at:
            return False
        try:
            return now >= datetime.fromisoformat(run_at)
        except ValueError:
            return False
    if kind == "daily":
        time_of_day = rec.get("time_of_day") or ""
        try:
            hh, mm = (int(x) for x in time_of_day.split(":")[:2])
        except (ValueError, AttributeError):
            return False
        days = rec.get("days_of_week")
        if days and now.weekday() not in days:
            return False
        if (now.hour, now.minute) < (hh, mm):
            return False
        if last_run_at:
            try:
                if datetime.fromisoformat(last_run_at).date() == now.date():
                    return False   # already ran today
            except ValueError:
                pass
        return True
    return False


def _run_due_schedules():
    """Poll tick: launch every enabled, due job through the SAME helpers a manual click uses
    (_do_launch_batch / _do_launch_crash_sweep) — a scheduled run gets the exact same
    validation and safety gates (accounts required, MAX_STAKE cap, etc.) a human click does.
    A job due while a run is already in progress is left alone (not consumed) so it fires on
    the next tick once the current run ends — the same single global run slot a manual click
    competes for."""
    schedules = load_schedules()
    if not schedules:
        return
    now = datetime.now()
    changed = False
    for job in schedules:
        if not job.get("enabled") or not _schedule_is_due(job, now):
            continue
        if _workers_alive():
            continue   # busy — stays due, retried next tick
        payload = job.get("payload") or {}
        try:
            if job.get("type") == "slot":
                result, status = _do_launch_batch(
                    payload.get("input_path"), payload.get("brand", "betway"),
                    payload.get("region", "ZA"), payload.get("accounts") or [],
                    bool(payload.get("headless")), payload.get("tests") or ["dsc"],
                    parallel=payload.get("parallel"))
            elif job.get("type") == "crash":
                result, status = _do_launch_crash_sweep(
                    payload.get("picks"), payload.get("username", ""), payload.get("password", ""),
                    payload.get("brand", "betway"), payload.get("region", "ZA"),
                    bool(payload.get("headless")), bool(payload.get("live")),
                    payload.get("bet"), payload.get("target"))
            else:
                result, status = {"message": f"unknown schedule type {job.get('type')!r}"}, 400
        except Exception as e:
            result, status = {"message": str(e)}, 500
        job["last_run_at"] = now.isoformat()
        job["last_run_id"] = result.get("run_id")
        job["last_status"] = "ok" if status == 200 else f"error: {result.get('message', status)}"
        if (job.get("recurrence") or {}).get("kind") == "once":
            job["enabled"] = False   # one-shot: done, kept in the list until deleted
        changed = True
    if changed:
        save_schedules(schedules)


def _schedule_loop():
    """Daemon thread started once in __main__ (see bottom of file) — distinct from the
    per-run fleet scheduler in _start_fleet, this one lives for the app's whole lifetime."""
    while True:
        try:
            _run_due_schedules()
        except Exception as e:
            print(f"[schedule] loop error: {e}")
        time.sleep(20)


# ─── Routes ──────────────────────────────────────────────────────

@app.route('/')
def index():
    return render_template('index.html')


@app.route('/api/accounts', methods=['GET'])
def get_accounts():
    return jsonify(load_accounts())


@app.route('/api/accounts', methods=['POST'])
def add_account():
    data = request.json
    accounts = load_accounts()
    # An account is unique per (brand, region, username) — same number can exist
    # across brands/regions as a separate login.
    brand = data.get("brand", "betway")
    region = data.get("region", "ZA")
    username = data.get("username", "")
    accounts = [a for a in accounts if not (
        a.get("username") == username
        and a.get("brand", "betway") == brand
        and a.get("region", "ZA") == region
    )]
    accounts.append({
        "label": data.get("label", ""),
        "username": username,
        "password": data.get("password", ""),
        "brand": brand,
        "region": region,
        "added": datetime.now().isoformat()
    })
    save_accounts(accounts)
    return jsonify({"status": "saved", "count": len(accounts)})


@app.route('/api/accounts/<username>', methods=['DELETE'])
def delete_account(username):
    brand = request.args.get("brand")
    region = request.args.get("region")
    accounts = load_accounts()

    def matches(a):
        if a.get("username") != username:
            return False
        if brand is not None and a.get("brand", "betway") != brand:
            return False
        if region is not None and a.get("region", "ZA") != region:
            return False
        return True

    accounts = [a for a in accounts if not matches(a)]
    save_accounts(accounts)
    return jsonify({"status": "deleted"})


@app.route('/api/schedules', methods=['POST'])
def create_schedule():
    """Create a scheduled job — the SAME payload /launch-batch (slot) or /launch-crash-sweep
    (crash) expects, plus a `recurrence`. Slot jobs are multipart (excel file +
    brand/region/accounts JSON/tests JSON/headless/recurrence JSON — same fields the manual
    batch upload sends, the excel is copied into SCHEDULED_UPLOADS_DIR so it survives past
    this request); crash jobs are a plain JSON body (picks/username/password/brand/region/
    headless/live/bet/target/recurrence). `recurrence`: {"kind":"once","run_at":"<ISO
    datetime>"} or {"kind":"daily","time_of_day":"HH:MM","days_of_week":[0-6] (optional,
    default every day, 0=Monday)}. Validated up front so a broken schedule is rejected at
    creation, not silently every day it tries to fire."""
    job_id = f"sch_{datetime.now():%Y%m%d_%H%M%S}_{uuid.uuid4().hex[:6]}"
    is_multipart = bool(request.content_type) and request.content_type.startswith("multipart/form-data")

    if is_multipart:
        job_type = "slot"
        f = request.files.get('excel')
        if not f:
            return jsonify({"status": "error", "message": "No Excel file uploaded"}), 400
        brand = request.form.get('brand', 'betway')
        region = request.form.get('region', 'ZA')
        headless = request.form.get('headless') == 'true'
        try:
            accounts = json.loads(request.form.get('accounts') or '[]')
        except ValueError:
            accounts = []
        accounts = [a for a in accounts if a.get("username") and a.get("password")]
        if not accounts:
            return jsonify({"status": "error",
                            "message": "Select at least one worker account"}), 400
        try:
            tests = json.loads(request.form.get('tests') or '["dsc"]') or ["dsc"]
        except ValueError:
            tests = ["dsc"]
        try:
            recurrence = json.loads(request.form.get('recurrence') or '{}')
        except ValueError:
            recurrence = {}
        fname = (f.filename or "").lower()
        if fname.endswith(".xls") and not fname.endswith(".xlsx"):
            return jsonify({"status": "error",
                            "message": "Legacy .xls isn't supported — save the sheet as .xlsx"}), 400
        input_path = os.path.join(SCHEDULED_UPLOADS_DIR, f"{job_id}_input.xlsx")
        if fname.endswith(".csv"):
            import pandas as pd
            csv_path = os.path.join(SCHEDULED_UPLOADS_DIR, f"{job_id}_input.csv")
            f.save(csv_path)
            try:
                pd.read_csv(csv_path).to_excel(input_path, index=False)
            except Exception as e:
                return jsonify({"status": "error", "message": f"Could not read the CSV: {e}"}), 400
        else:
            f.save(input_path)
        payload = {"input_path": input_path, "brand": brand, "region": region,
                   "accounts": accounts, "headless": headless, "tests": tests}
    else:
        data = request.get_json(force=True) or {}
        job_type = "crash"
        username, password = data.get('username', ''), data.get('password', '')
        picks = [p for p in (data.get('picks') or []) if (p or {}).get('name')]
        if not username or not password:
            return jsonify({"status": "error", "message": "Select an account first"}), 400
        if not picks:
            return jsonify({"status": "error", "message": "No crash games selected"}), 400
        recurrence = data.get('recurrence') or {}
        payload = {"picks": picks, "username": username, "password": password,
                   "brand": (data.get('brand') or 'betway').lower(),
                   "region": (data.get('region') or 'ZA').upper(),
                   "headless": bool(data.get('headless')), "live": bool(data.get('live')),
                   "bet": data.get('bet'), "target": data.get('target')}

    kind = recurrence.get("kind")
    if kind == "once":
        if not recurrence.get("run_at"):
            return jsonify({"status": "error", "message": "Provide a run_at date/time"}), 400
    elif kind == "daily":
        if not recurrence.get("time_of_day"):
            return jsonify({"status": "error", "message": "Provide a time_of_day (HH:MM)"}), 400
    else:
        return jsonify({"status": "error",
                        "message": "recurrence.kind must be 'once' or 'daily'"}), 400

    job = {"id": job_id, "type": job_type, "enabled": True, "recurrence": recurrence,
           "payload": payload, "created_at": datetime.now().isoformat(),
           "last_run_at": None, "last_run_id": None, "last_status": None}
    schedules = load_schedules()
    schedules.append(job)
    save_schedules(schedules)
    return jsonify({"status": "created", "job": job})


@app.route('/api/schedules', methods=['GET'])
def list_schedules():
    """Every scheduled job (slot + crash) — feeds the dashboard's Upcoming Schedules list."""
    return jsonify(load_schedules())


@app.route('/api/schedules/<job_id>', methods=['PATCH'])
def update_schedule(job_id):
    """Partial update — primarily {"enabled": bool} to pause/resume a job, but also accepts a
    replacement {"recurrence": {...}}."""
    data = request.get_json(force=True) or {}
    schedules = load_schedules()
    job = next((j for j in schedules if j.get("id") == job_id), None)
    if not job:
        return jsonify({"status": "error", "message": "Schedule not found"}), 404
    if "enabled" in data:
        job["enabled"] = bool(data["enabled"])
    if isinstance(data.get("recurrence"), dict):
        job["recurrence"] = data["recurrence"]
    save_schedules(schedules)
    return jsonify({"status": "updated", "job": job})


@app.route('/api/schedules/<job_id>', methods=['DELETE'])
def delete_schedule(job_id):
    """Remove a scheduled job; for a slot job also deletes its persisted upload copy."""
    schedules = load_schedules()
    job = next((j for j in schedules if j.get("id") == job_id), None)
    if not job:
        return jsonify({"status": "error", "message": "Schedule not found"}), 404
    if job.get("type") == "slot":
        input_path = (job.get("payload") or {}).get("input_path")
        if input_path and os.path.exists(input_path):
            try:
                os.remove(input_path)
            except OSError:
                pass
    schedules = [j for j in schedules if j.get("id") != job_id]
    save_schedules(schedules)
    return jsonify({"status": "deleted"})


@app.route('/api/results')
def get_results():
    if os.path.exists(RESULTS_FILE):
        with open(RESULTS_FILE, "r") as f:
            return jsonify(json.load(f))
    return jsonify([])


@app.route('/api/games/search', methods=['POST'])
def games_search():
    """Live typeahead: authenticate with the selected account (token cached) and return matching
    games for the query. Body: {q, username, password, brand, region}."""
    data = request.json or {}
    q = (data.get('q') or '').strip()
    username = data.get('username', '')
    password = data.get('password', '')
    brand = (data.get('brand') or 'betway').lower()
    region = (data.get('region') or 'ZA').upper()
    if len(q) < 2:
        return jsonify({"games": []})
    if not username or not password:
        return jsonify({"games": [], "error": "select an account to search"})
    try:
        token, err = _get_token(username, password, brand=brand, region=region)
        if not token:
            return jsonify({"games": [], "error": f"auth failed: {err}"})
        from modules.game_handler import GameHandler
        games = GameHandler().search_games_list(q, token, region=region, brand=brand, limit=10)
        return jsonify({"games": games})
    except Exception as e:
        print(f"[games_search] EXCEPTION q={q!r}: {e}")
        return jsonify({"games": [], "error": str(e)})


@app.route('/api/recordings')
def get_recordings():
    files = sorted(glob.glob(os.path.join(RECORDINGS_DIR, "*.webm")), key=os.path.getmtime, reverse=True)
    result = []
    for f in files[:10]:
        name = os.path.basename(f)
        size_mb = round(os.path.getsize(f) / (1024 * 1024), 1)
        mtime = datetime.fromtimestamp(os.path.getmtime(f)).strftime("%b %d, %H:%M")
        result.append({"name": name, "size": f"{size_mb} MB", "date": mtime})
    return jsonify(result)


@app.route('/recordings/<path:filename>')
def serve_recording(filename):
    return send_from_directory(RECORDINGS_DIR, filename)


@app.route('/api/screenshot')
def latest_screenshot():
    pngs = glob.glob(os.path.join(SCREENSHOTS_DIR, "*.png"))
    if not pngs:
        return "", 204
    latest = max(pngs, key=os.path.getmtime)
    return send_from_directory(SCREENSHOTS_DIR, os.path.basename(latest))


@app.route('/screenshots/<path:filename>')
def serve_screenshot(filename):
    return send_from_directory(SCREENSHOTS_DIR, filename)


@app.route('/runs/<run_id>/<path:filename>')
def serve_run_file(run_id, filename):
    """Serve any artifact (screenshots/, video/, logs/, results.json) from a run folder."""
    # Guard against path traversal in run_id.
    if not re.fullmatch(r"[A-Za-z0-9._-]+", run_id or ""):
        abort(404)
    return send_from_directory(os.path.join(RUNS_DIR, run_id), filename)


@app.route('/dsc-report/latest')
def dsc_report_latest():
    """Download the most recent DSC Excel report — batch (DSC_Report_*), auto-sweep (DSC_Sweep_*),
    or CLI sweep (DSC_Auto_*). Input sheets (*_input.xlsx) are not reports."""
    reports = [p for p in glob.glob(os.path.join(RUNS_DIR, "DSC_*.xlsx"))
               if not p.endswith("_input.xlsx")]
    if not reports:
        return "No DSC report yet — run a sanity check first.", 404
    latest = max(reports, key=os.path.getmtime)
    return send_from_directory(RUNS_DIR, os.path.basename(latest), as_attachment=True)


@app.route('/api/runs')
def list_runs():
    """Recent runs (newest first) — feeds the homepage history panel."""
    if os.path.exists(RUNS_INDEX):
        with open(RUNS_INDEX, "r", encoding="utf-8") as f:
            return jsonify(json.load(f))
    return jsonify([])


def _batch_report_path(run_id):
    """The Excel report belonging to a batch run: from its index entry when recorded,
    else derived from the run_id timestamp (runs/<YYYYMMDD_HHMMSS>_DSC ↔
    DSC_Report_<YYYY-MM-DD_HHMMSS>.xlsx)."""
    try:
        if os.path.exists(RUNS_INDEX):
            with open(RUNS_INDEX, "r", encoding="utf-8") as f:
                for e in json.load(f):
                    if e.get("run_id") == run_id and e.get("report"):
                        return os.path.join(RUNS_DIR, e["report"])
    except Exception:
        pass
    m = re.match(r"^(\d{4})(\d{2})(\d{2})_(\d{6})_DSC$", run_id or "")
    if m:
        p = os.path.join(RUNS_DIR,
                         f"DSC_Report_{m.group(1)}-{m.group(2)}-{m.group(3)}_{m.group(4)}.xlsx")
        if os.path.exists(p):
            return p
    return None


def _report_rows(report_path):
    """The report sheet as JSON rows (blank/filler rows skipped) for the homepage
    validation table."""
    from openpyxl import load_workbook
    from modules import dsc_report as dr
    wb = load_workbook(report_path)
    ws, hdr, cols = dr._pick_sheet(wb)
    idxs = {c: dr._col(cols, c) for c in dr.DSC_COLUMNS}
    out = []
    for r in range(hdr + 1, ws.max_row + 1):
        name = ws.cell(row=r, column=idxs["Game Name"]).value if idxs["Game Name"] else None
        if not str(name or "").strip():
            continue
        out.append({c: (ws.cell(row=r, column=i).value if i else "") for c, i in idxs.items()})
    wb.close()
    return out


def _crash_sweep_rows(run_id):
    """Aggregate a crash sweep's per-game results.json files into one table. Unlike the slot
    batch/auto-sweep paths, a crash sweep has NO single shared report file — /launch-crash-sweep
    runs one crash_auto.py subprocess per game, each writing its OWN results.json into
    runs/<batch_id>/<NN_slug>/ (see app.py's launch_crash_sweep). So this reads those folders
    directly off disk instead of parsing an Excel report — works even mid-run (a game still in
    progress just has no results.json yet and is skipped this pass, not shown as failed)."""
    batch_dir = os.path.join(RUNS_DIR, run_id)
    rows = []
    for d in sorted(glob.glob(os.path.join(batch_dir, "*"))):
        folder = os.path.basename(d)
        rj = os.path.join(d, "results.json")
        if not os.path.isdir(d) or not os.path.exists(rj):
            continue
        name = re.sub(r"^\d+_", "", folder).replace("-", " ").strip() or folder
        try:
            with open(rj, encoding="utf-8") as f:
                data = json.load(f)
            summary = data.get("summary", {})
            rows.append({"folder": folder, "name": name,
                        "total": summary.get("total", 0), "passed": summary.get("passed", 0),
                        "failed": summary.get("failed", 0), "skipped": summary.get("skipped", 0)})
        except Exception as e:
            rows.append({"folder": folder, "name": name, "error": str(e)})
    return rows


CRASH_REPORT_COLUMNS = ["Sr. No.", "Provider", "Game Name", "Game Type",
                       "Launch", "Bet Placed", "Tlogs", "Remark", "Evidence"]


def _crash_report_rows(run_id):
    """Same per-game folders _crash_sweep_rows reads, mapped onto the DSC-style column set
    (see modules/dsc_report.py DSC_COLUMNS) plus Game Type — so crash results sit in the same
    shape as the slot sheet's, importable into the team's master tracker.

    Launch: Pass if 'Crash UI controls detected' (TEST 1) passed — i.e. the game genuinely
    loaded and rendered its UI. Deliberately NOT gated on TEST 2 (round-lifecycle observed)
    too: TEST 2 can fail on a perfectly-launched game just because the round we happened to
    watch ran long and our observation window timed out — that's a lifecycle-OBSERVATION gap,
    not evidence the game never launched, and conflating the two mis-marked real launches as
    "Did not launch" (confirmed 2026-07-28 against a real FlyX Party run).
    A run can also abort BEFORE TEST 1 ever runs (crash_auto.py's _abort_session — dead/expired
    session, or the phase drifted unreadable right after startup). That path writes a single
    result named "Crash game session is live" instead of the normal TEST 1-4 battery; handled
    explicitly below so its own specific reason surfaces instead of falling through to a vague
    generic message.
    Bet Placed: read from crash_auto.py's own "Place bet during betting window = 1 place-bet
    request" result — under --live this actually places the wager and passed=True/False reports
    whether it went through; under dry-run (or when the run aborted before reaching that test)
    it's recorded with passed=None, which maps to "NA" here. (Fixed 2026-08-11 — this used to
    hardcode "NA" and label every launched game "Dry-run only — no wagers placed" unconditionally,
    a leftover from before /launch-crash-sweep supported --live; that made every live crash sweep's
    report claim no money was wagered even when it genuinely was.)
    """
    batch_dir = os.path.join(RUNS_DIR, run_id)
    rows = []
    for i, d in enumerate(sorted(glob.glob(os.path.join(batch_dir, "*"))), 1):
        folder = os.path.basename(d)
        rj = os.path.join(d, "results.json")
        if not os.path.isdir(d) or not os.path.exists(rj):
            continue
        pick = {}
        pj = os.path.join(d, "pick.json")
        if os.path.exists(pj):
            try:
                with open(pj, encoding="utf-8") as f:
                    pick = json.load(f)
            except Exception:
                pick = {}
        name = pick.get("name") or re.sub(r"^\d+_", "", folder).replace("-", " ").strip() or folder
        try:
            with open(rj, encoding="utf-8") as f:
                data = json.load(f)
            results = {r.get("name"): r for r in data.get("results", [])}
            abort = results.get("Crash game session is live")
            bet_test = results.get("Place bet during betting window = 1 place-bet request")
            if abort is not None:
                launched = bool(abort.get("passed"))
                remark = abort.get("details") or "Session did not stay live long enough to test"
                bet_placed = "NA"
            else:
                t1 = results.get("Crash UI controls detected", {})
                launched = bool(t1.get("passed"))
                if bet_test is None or bet_test.get("passed") is None:
                    bet_placed = "NA"
                    remark = (bet_test.get("details") if bet_test else None) or \
                        ("Dry-run — no wager attempted" if launched else
                         (t1.get("details") or "Controls not detected"))
                else:
                    bet_placed = "Pass" if bet_test.get("passed") else "Fail"
                    remark = bet_test.get("details") or \
                        ("Wager placed" if bet_placed == "Pass" else "Wager attempt failed")
                    # Cross-check against the balance-delta test (2026-08-11): the bet-request
                    # test can false-negative when a provider's real wager confirmation travels
                    # over a WS frame the network monitor classifies as idle (confirmed live
                    # against "Trader" — bet test said Fail/requests=0, but balance genuinely
                    # dropped by the stake amount). A non-zero balance delta is stronger evidence
                    # of a real wager than the request-count heuristic, so it overrides here.
                    if bet_placed != "Pass":
                        bal_test = results.get("Balance updates correctly (-bet, +payout on cash out)")
                        m = re.search(r"delta\s+([+-]?\d+\.?\d*)", (bal_test or {}).get("details") or "")
                        if m and abs(float(m.group(1))) > 0.001:
                            bet_placed = "Pass"
                            remark = f"Wager confirmed via balance drop (bet-request test missed it): {bal_test['details']}"
            rows.append({"Sr. No.": i, "Provider": pick.get("provider") or "",
                        "Game Name": name, "Game Type": "Crash Games",
                        "Launch": "Pass" if launched else "Fail",
                        "Bet Placed": bet_placed, "Tlogs": "NA", "Remark": remark,
                        "Evidence": f"{run_id}/{folder}"})
        except Exception as e:
            rows.append({"Sr. No.": i, "Provider": pick.get("provider") or "", "Game Name": name,
                        "Game Type": "Crash Games", "Launch": "Fail", "Bet Placed": "NA",
                        "Tlogs": "NA", "Remark": f"Unreadable results.json: {e}",
                        "Evidence": f"{run_id}/{folder}"})
    return rows


def _write_crash_report_xlsx(run_id, rows):
    from openpyxl import Workbook
    from openpyxl.styles import Font
    path = os.path.join(RUNS_DIR, f"Crash_Report_{run_id}.xlsx")
    wb = Workbook()
    ws = wb.active
    ws.title = "Crash"
    ws.append(CRASH_REPORT_COLUMNS)
    widths = {"Sr. No.": 8, "Provider": 22, "Game Name": 30, "Game Type": 14,
             "Launch": 10, "Bet Placed": 12, "Tlogs": 10, "Remark": 46, "Evidence": 34}
    for cell in ws[1]:
        cell.font = Font(bold=True)
        ws.column_dimensions[cell.column_letter].width = widths.get(cell.value, 14)
    for row in rows:
        ws.append([row.get(c, "") for c in CRASH_REPORT_COLUMNS])
    wb.save(path)
    return path


@app.route('/api/crash-report/<run_id>')
def api_crash_report(run_id):
    """Build (or rebuild) the Excel report for a crash sweep on demand, in the same column
    shape the slot DSC sheet uses (see _crash_report_rows). Regenerated fresh each call — cheap
    (a handful of small JSON files) and always reflects the latest results on disk, including a
    sweep that's still running. run_id='latest' is handled by api_crash_report_latest below —
    Flask matches the static rule first regardless of declaration order, so this dynamic route
    never actually sees the literal string 'latest'."""
    if not re.fullmatch(r"[A-Za-z0-9._-]+", run_id or "") or not run_id.endswith("_crash_sweep"):
        abort(404)
    rows = _crash_report_rows(run_id)
    if not rows:
        return jsonify({"error": "no saved results for this run yet"}), 404
    path = _write_crash_report_xlsx(run_id, rows)
    return send_from_directory(RUNS_DIR, os.path.basename(path), as_attachment=True)


@app.route('/api/crash-report/latest')
def api_crash_report_latest():
    """One-click download of the most recent crash sweep's report — the crash-vertical
    equivalent of /dsc-report/latest, for the persistent link in the Crash Sweep card (no need
    to open Recent Runs first, matching how slots expose 'Download DSC report')."""
    index = []
    if os.path.exists(RUNS_INDEX):
        with open(RUNS_INDEX, "r", encoding="utf-8") as fh:
            index = json.load(fh)
    latest = next((r for r in index if r.get("type") == "crash-sweep"), None)
    if not latest:
        return "No crash sweep run yet — run one first.", 404
    return api_crash_report(latest["run_id"])


@app.route('/api/report/<run_id>')
def api_report(run_id):
    """Persisted results for a past run, so the report survives page refreshes.
    Single runs: the run folder's results.json (same payload the live stream emits).
    Batch runs: the Excel report parsed into rows + a tally.
    Crash sweeps: no Excel report — aggregated on the fly from each game's own results.json."""
    if not re.fullmatch(r"[A-Za-z0-9._-]+", run_id or ""):
        abort(404)
    # Batch AND auto-sweep runs are Excel-report-backed (many rows); render them as a table.
    if run_id.endswith("_DSC") or run_id.endswith("_DSCsweep"):
        rep = _batch_report_path(run_id)
        if not rep:
            return jsonify({"error": "no report found for this run"}), 404
        return jsonify({"batch": True, "run_id": run_id,
                        "report": os.path.basename(rep),
                        "summary": _report_summary(rep),
                        "rows": _report_rows(rep)})
    if run_id.endswith("_crash_sweep"):
        rows = _crash_sweep_rows(run_id)
        if not rows:
            return jsonify({"error": "no saved results for this run yet"}), 404
        return jsonify({"batch": True, "crash": True, "run_id": run_id, "rows": rows})
    rj = os.path.join(RUNS_DIR, run_id, "results.json")
    if os.path.exists(rj):
        with open(rj, "r", encoding="utf-8") as f:
            return jsonify(json.load(f))
    return jsonify({"error": "no saved results for this run"}), 404


@app.route('/dsc-report/file/<name>')
def dsc_report_file(name):
    """Download a SPECIFIC report (the /latest route only serves the newest). Covers batch
    (DSC_Report_*), auto-sweep (DSC_Sweep_*) and CLI-sweep (DSC_Auto_*) reports."""
    if not re.fullmatch(r"DSC_[A-Za-z0-9._-]+\.xlsx", name or "") or name.endswith("_input.xlsx"):
        abort(404)
    return send_from_directory(RUNS_DIR, name, as_attachment=True)


# ─── Tlogs validation: verify recorded spins against the site's transaction history ───
@app.route('/tlogs-reports')
def tlogs_reports():
    """Bet-records files (newest first) the validator can be pointed at."""
    out = []
    for p in sorted(glob.glob(os.path.join(RUNS_DIR, "*_records.jsonl")),
                    key=os.path.getmtime, reverse=True)[:20]:
        try:
            with open(p, encoding="utf-8") as f:
                n = sum(1 for line in f if line.strip())
        except OSError:
            n = 0
        out.append({"name": os.path.basename(p), "records": n,
                    "age_min": int((time.time() - os.path.getmtime(p)) / 60),
                    "validated": os.path.exists(p.replace("_records.jsonl", "_validation.json"))})
    return jsonify(out)


@app.route('/validate-tlogs', methods=['POST'])
def validate_tlogs():
    """Launch tlogs_validate.py as a worker; output streams through /stream like any run."""
    global CURRENT_RUN
    if _workers_alive():
        return jsonify({"status": "error", "message": "A run is already in progress"}), 409
    data = request.json or {}
    name = os.path.basename(data.get('records') or '')
    path = os.path.join(RUNS_DIR, name)
    if not name.endswith('_records.jsonl') or not os.path.exists(path):
        return jsonify({"status": "error", "message": "Records file not found"}), 400
    cmd = _worker_cmd("tlogs_validate.py", "--records", path)
    if data.get('headed'):
        cmd.append("--headed")
    CURRENT_RUN = {"game": f"Tlogs validation · {name}", "run_id": name,
                   "start_time": datetime.now().isoformat(), "status": "running"}
    _start_workers([("TLOGS", cmd)])
    return jsonify({"status": "started"})


@app.route('/tlogs-validation/latest')
def tlogs_validation_latest():
    """The most recent validation result (JSON written next to the records file)."""
    files = glob.glob(os.path.join(RUNS_DIR, "*_validation.json"))
    if not files:
        return jsonify({"error": "no validation yet"}), 404
    latest = max(files, key=os.path.getmtime)
    with open(latest, encoding="utf-8") as f:
        return jsonify(json.load(f))


@app.route('/api/log-history')
def log_history():
    """The current/most recent run's streamed lines — lets the UI repopulate the
    log pane after a refresh instead of showing a blank terminal."""
    return jsonify({"lines": list(LOG_HISTORY),
                    "game": CURRENT_RUN.get("game"),
                    "status": CURRENT_RUN.get("status")})


@app.route('/api/status')
def status():
    global CURRENT_RUN
    if _workers_alive():
        CURRENT_RUN["status"] = "running"
    elif CURRENT_RUN["status"] == "running":
        CURRENT_RUN["status"] = "done"
    return jsonify(CURRENT_RUN)


@app.route('/launch', methods=['POST'])
def launch():
    global CURRENT_RUN
    data = request.json

    game_type = data.get('game_type', 'slot')
    game_name = data.get('game', '')
    username = data.get('username', '')
    password = data.get('password', '')
    is_mobile = data.get('mobile', False)
    default_bet = data.get('default_bet', '')
    min_bet = data.get('min_bet', '')
    brand = data.get('brand', 'betway')
    region = data.get('region', 'ZA')
    is_headless = data.get('headless', False)
    tests = data.get('tests')   # optional list of gated checks; None/absent = run all
    crash_url = data.get('url', '')

    # Each launch gets its own folder: runs/<timestamp>_<game>/ holding screenshots/, video/,
    # logs/, results.json — so runs never overwrite each other.
    started = datetime.now()

    if game_type == 'crash':
        # LIVE ONLY — the dry-run path was deliberately removed from this card (2026-07-29,
        # explicit request) so every run through it places a real wager, gated by
        # crash_auto.py --live + config_env.MAX_STAKE, PLUS the dashboard's own confirmation
        # checkbox. Stake is OPTIONAL (2026-08-10): crash_auto.py resolves each game's own
        # minimum bet automatically (catalog minBetAmount, or in-game floor detection for a
        # raw URL with no catalog lookup) — a typed value here is only an override/cap, still
        # re-validated server-side since it spends real money.
        import config_env
        live_target = (data.get('target') or '').strip()
        try:
            live_bet = float(data.get('bet') or 0)
        except (TypeError, ValueError):
            live_bet = 0
        if live_bet < 0:
            return jsonify({"status": "error", "message": "Stake override cannot be negative"}), 400
        if live_bet > config_env.MAX_STAKE:
            return jsonify({"status": "error",
                            "message": f"Stake {live_bet:g} exceeds the safety cap "
                                       f"{config_env.MAX_STAKE:g}"}), 400
        label = game_name or ("Crash URL" if crash_url else "Crash game")
        run_id = f"{started:%Y%m%d_%H%M%S}_crash_{_slug(label)}"
        run_dir = os.path.join(RUNS_DIR, run_id)
        os.makedirs(run_dir, exist_ok=True)
        cmd = _worker_cmd("crash_auto.py")
        if crash_url:
            cmd.append(crash_url)
        elif game_name:
            if not username or not password:
                return jsonify({"status": "error",
                                "message": "Crash-by-name needs an account (username + password), "
                                           "or paste a direct launch URL"}), 400
            cmd += ["--game", game_name, "--username", username, "--password", password]
        else:
            return jsonify({"status": "error",
                            "message": "Provide a crash launch URL or a game name"}), 400
        cmd += ["--live"]
        if live_bet > 0:
            cmd += ["--bet", str(live_bet)]
        if live_target:
            cmd += ["--target", live_target]
        cmd += ["--run-dir", run_dir, "--brand", brand, "--region", region]
        if is_mobile:
            cmd.append("--mobile")
        if is_headless:
            cmd.append("--headless")
        game_name = label   # label for CURRENT_RUN + the runs index below
    else:
        if not game_name or not username or not password:
            return jsonify({"status": "error", "message": "Missing fields"}), 400

        run_id = f"{started:%Y%m%d_%H%M%S}_{_slug(game_name)}"
        run_dir = os.path.join(RUNS_DIR, run_id)
        os.makedirs(run_dir, exist_ok=True)

        cmd = _worker_cmd("test_spin_button.py",
               "--game", game_name,
               "--username", username,
               "--password", password,
               "--brand", brand,
               "--region", region,
               "--run-dir", run_dir)
        if is_mobile:
            cmd.append("--mobile")
        if is_headless:
            cmd.append("--headless")
        if default_bet:
            cmd.extend(["--default-bet", default_bet])
        if min_bet:
            cmd.extend(["--min-bet", min_bet])
        # Pass --tests whenever the UI sent a selection list. Empty list => the user unchecked every
        # check (including core) => "none". Absent (None) => run everything.
        if isinstance(tests, list):
            cmd.extend(["--tests", ",".join(str(t) for t in tests) if tests else "none"])

    CURRENT_RUN = {
        "game": game_name,
        "run_id": run_id,
        "start_time": started.isoformat(),
        "status": "running"
    }
    # Prepend a record to the runs index (newest first).
    try:
        index = []
        if os.path.exists(RUNS_INDEX):
            with open(RUNS_INDEX, "r", encoding="utf-8") as f:
                index = json.load(f)
        # Crash's single-game card is LIVE-only now (2026-07-29) and writes real bet records
        # (see run_crash_tests' dsc_report.append_record call) — tag it distinctly so the
        # dashboard's Validate button can find it, without touching slot's "single" type at all.
        index.insert(0, {"run_id": run_id, "game": game_name, "brand": brand,
                         "region": region, "started_at": started.isoformat(),
                         "type": "crash-live" if game_type == "crash" else "single"})
        with open(RUNS_INDEX, "w", encoding="utf-8") as f:
            json.dump(index[:200], f, indent=2)
    except Exception as e:
        print(f"[runs] index write failed: {e}")

    _start_workers([("", cmd)])

    return jsonify({"status": "started"})


def _do_launch_batch(input_path, brand, region, accounts, headless, tests, parallel=None):
    """Core of a slot batch launch: shard `input_path` across `accounts`, seed a shared DSC
    report, spawn one worker per shard, register the run. Factored out of the /launch-batch
    route (2026-08-10) so the scheduler can trigger the EXACT same launch a manual click does
    — same validation, same worker commands — just called with a stored payload instead of a
    fresh request. `input_path` must already point at a saved .xlsx (the route saves the
    upload before calling this; the scheduler points at its persisted per-job copy).

    `parallel` caps how many of the (up to one-per-account) workers run at once — worker
    accounts still queue and start in waves as slots free (see _start_workers), instead of
    every browser opening at the same instant. Default/None -> 4, same default+[1,8] clamp the
    auto-sweep fleet already uses; a big batch (e.g. 30+ accounts) opening that many real
    Chromium instances simultaneously starves CPU/RAM AND floods the single shared Gemini
    vision API key with concurrent calls, so the cap applies even when the caller passes nothing.

    Returns (payload_dict, http_status) — same shape the route used to jsonify directly."""
    global CURRENT_RUN
    accounts = [a for a in (accounts or []) if a.get("username") and a.get("password")]
    if not accounts:
        return {"status": "error", "message": "Select at least one worker account"}, 400
    tests = tests or ["dsc"]
    try:
        parallel = max(1, min(int(parallel or 4), 8))
    except (TypeError, ValueError):
        parallel = 4

    started = datetime.now()
    batch_id = f"{started:%Y%m%d_%H%M%S}_DSC"
    batch_dir = os.path.join(RUNS_DIR, batch_id)
    os.makedirs(batch_dir, exist_ok=True)

    from modules import dsc_report
    try:
        shards, total = dsc_report.shard_excel(input_path, batch_dir, len(accounts))
    except Exception as e:
        return {"status": "error", "message": f"Could not read the sheet: {e}"}, 400

    report = None
    if "dsc" in tests:
        # Fresh per-batch report (not the shared daily file): seeded from the FULL input
        # so row order survives sharding. Lives directly in runs/ so /dsc-report/latest
        # finds it. Workers' own ensure_report calls see it exists and leave it alone.
        report = os.path.join(RUNS_DIR, f"DSC_Report_{started:%Y-%m-%d_%H%M%S}.xlsx")
        dsc_report.ensure_report(report, seed_from=input_path)

    cmds = []
    for k, (shard, acc) in enumerate(zip(shards, accounts), 1):
        cmd = _worker_cmd("test_spin_button.py",
               "--excel", shard,
               "--username", acc["username"],
               "--password", acc["password"],
               "--brand", brand,
               "--region", region,
               "--run-dir", os.path.join(batch_dir, f"w{k}"),
               "--tests", ",".join(str(t) for t in tests))
        if headless:
            cmd.append("--headless")
        if report:
            cmd.extend(["--dsc-report", report])
        # Cascade the worker windows (~36px steps) so every title bar stays clickable —
        # users resize/snap stacked windows to see them, and a RESIZE breaks coordinates.
        # Stepped modulo `parallel` (mirrors the auto-sweep fleet) since only `parallel`
        # workers are ever actually on screen at once now — without the modulo a big batch
        # would cascade windows off the bottom of the screen for waves that never overlap.
        env = os.environ.copy()
        step = 36 * ((k - 1) % parallel)
        env["GAMEGUARD_WINDOW_POS"] = f"{step},{step}"
        cmds.append((f"W{k}", cmd, env))

    batch = {"report": report, "total": total, "workers": len(cmds), "run_id": batch_id,
             "parallel": parallel}
    CURRENT_RUN = {"game": f"Batch · {total} games", "run_id": batch_id,
                   "start_time": started.isoformat(), "status": "running", "batch": batch}
    try:
        index = []
        if os.path.exists(RUNS_INDEX):
            with open(RUNS_INDEX, "r", encoding="utf-8") as fh:
                index = json.load(fh)
        index.insert(0, {"run_id": batch_id, "game": f"Batch · {total} games",
                         "brand": brand, "region": region,
                         "started_at": started.isoformat(), "workers": len(cmds),
                         "type": "batch",
                         "report": os.path.basename(report) if report else None})
        with open(RUNS_INDEX, "w", encoding="utf-8") as fh:
            json.dump(index[:200], fh, indent=2)
    except Exception as e:
        print(f"[runs] index write failed: {e}")

    _start_workers(cmds, batch=batch, max_parallel=parallel)

    return {"status": "started", "total": total, "workers": len(cmds), "parallel": parallel,
            "run_id": batch_id, "report": os.path.basename(report) if report else None}, 200


@app.route('/launch-batch', methods=['POST'])
def launch_batch():
    """Batch DSC sweep: an uploaded Excel of games, split round-robin across one browser
    worker per selected account. All workers fill ONE shared report (file-locked) seeded
    from the input sheet, so row order is preserved and a partial sweep is readable.
    Multipart form: excel (file), brand, region, accounts (JSON [{username,password}]),
    tests (JSON list, default ["dsc"]), parallel (optional int, default 4, clamped [1,8] —
    how many worker browsers run at once; the rest queue). Thin wrapper around
    _do_launch_batch — this route's only job is turning the multipart upload into a saved
    input_path."""
    f = request.files.get('excel')
    if not f:
        return jsonify({"status": "error", "message": "No Excel file uploaded"}), 400
    brand = request.form.get('brand', 'betway')
    region = request.form.get('region', 'ZA')
    is_headless = request.form.get('headless') == 'true'
    try:
        accounts = json.loads(request.form.get('accounts') or '[]')
    except ValueError:
        accounts = []
    accounts = [a for a in accounts if a.get("username") and a.get("password")]
    if not accounts:
        return jsonify({"status": "error",
                        "message": "Select at least one worker account"}), 400
    try:
        tests = json.loads(request.form.get('tests') or '["dsc"]') or ["dsc"]
    except ValueError:
        tests = ["dsc"]
    try:
        parallel = int(request.form.get('parallel')) if request.form.get('parallel') else None
    except ValueError:
        parallel = None

    fname = (f.filename or "").lower()
    if fname.endswith(".xls") and not fname.endswith(".xlsx"):
        return jsonify({"status": "error",
                        "message": "Legacy .xls isn't supported — save the sheet as .xlsx"}), 400
    # Staged outside any particular run's own folder — _do_launch_batch mints its OWN
    # batch_id/run folder (shared with the scheduler, which calls it with a persisted upload
    # of its own), so the raw upload has nowhere fixed to land ahead of that.
    uploads_dir = os.path.join(RUNS_DIR, "_uploads")
    os.makedirs(uploads_dir, exist_ok=True)
    stamp = datetime.now().strftime("%Y%m%d_%H%M%S_%f")
    input_path = os.path.join(uploads_dir, f"{stamp}_input.xlsx")
    if fname.endswith(".csv"):
        # The dashboard file picker accepts .csv but openpyxl only reads xlsx — convert.
        import pandas as pd
        csv_path = os.path.join(uploads_dir, f"{stamp}_input.csv")
        f.save(csv_path)
        try:
            pd.read_csv(csv_path).to_excel(input_path, index=False)
        except Exception as e:
            return jsonify({"status": "error", "message": f"Could not read the CSV: {e}"}), 400
    else:
        f.save(input_path)

    payload, status = _do_launch_batch(input_path, brand, region, accounts, is_headless, tests,
                                       parallel=parallel)
    return jsonify(payload), status


def _sweep_targets(data):
    """Validate the {targets:[{brand,region,username,password,label?}]} a scope resolves to."""
    return [t for t in (data.get('targets') or [])
            if t.get('brand') and t.get('region') and t.get('username') and t.get('password')]


@app.route('/api/configured-pairs')
def api_configured_pairs():
    """The (brand, region) pairs that are actually testable — drives the scope→region resolution
    in the UI (which regions a scope covers, so it can auto-map saved accounts to them)."""
    from modules.utils import configured_pairs
    return jsonify({"pairs": [{"brand": b, "region": r} for (b, r) in configured_pairs()]})


@app.route('/api/sweep-plan', methods=['POST'])
def api_sweep_plan():
    """Preview the rotation plan grouped PER (brand, region) — NO spins, no money. Shows exactly
    which game each provider would get in each region so the plan is known before launch.
    JSON body: targets[{brand,region,username,password,label?}], limit, providers[]."""
    import dsc_sweep
    data = request.get_json(force=True) or {}
    targets = _sweep_targets(data)
    if not targets:
        return jsonify({"status": "error", "message": "No account resolved for the chosen scope — "
                        "add a saved account for each region you want to test"}), 400
    try:
        groups = dsc_sweep.plan_sweep(targets, data.get('providers') or None,
                                      data.get('limit') or None,
                                      one_per_region=bool(data.get('one_per_region')))
    except Exception as e:
        return jsonify({"status": "error", "message": f"Planning failed: {e}"}), 400
    out = [{"brand": g["brand"], "region": g["region"],
            "label": (g["accounts"][0].get("label") if g.get("accounts") else None),
            "account": (g["accounts"][0].get("username") if g.get("accounts") else None),
            "browsers": len(g.get("accounts", [])),
            "dropped": g.get("dropped", 0),
            "error": g.get("error"),
            "picks": [{"provider": p["provider"], "game": p["name"], "id": p["id"]}
                      for p in g.get("picks", [])]}
           for g in groups]
    total = sum(len(g["picks"]) for g in out)
    return jsonify({"groups": out, "total": total})


@app.route('/launch-sweep', methods=['POST'])
def launch_sweep():
    """Auto-DSC across one or many (brand, region)s. For each region: enumerate every provider,
    pick ONE untested game each (7-day rotation), then run the DSC — its games sharded across the
    region's selected accounts (one browser each). The whole fleet runs at most `parallel` browsers
    at a time. Results fill a per-region report and are recorded to the prod DB on completion
    (unless do_not_record). JSON body: targets[{brand,region,username,password,label?}], parallel,
    headless, do_not_record, limit, providers[]."""
    global CURRENT_RUN
    data = request.get_json(force=True) or {}
    is_headless = bool(data.get('headless'))
    do_not_record = bool(data.get('do_not_record'))
    limit = data.get('limit') or None
    only = data.get('providers') or None
    try:
        parallel = max(1, min(int(data.get('parallel') or 4), 8))
    except (TypeError, ValueError):
        parallel = 4
    targets = _sweep_targets(data)
    if not targets:
        return jsonify({"status": "error", "message": "No account resolved for the chosen scope"}), 400

    from modules import dsc_report
    import dsc_sweep

    try:
        groups = dsc_sweep.plan_sweep(targets, only, limit, max_parallel=parallel,
                                      one_per_region=bool(data.get('one_per_region')))
    except Exception as e:
        return jsonify({"status": "error", "message": f"Provider planning failed: {e}"}), 400
    planned = [g for g in groups if g.get("picks")]
    if not planned:
        errs = "; ".join(f"{g['brand']} {g['region']}: {g['error']}"
                         for g in groups if g.get("error"))
        filt = f" No provider matched your filter '{', '.join(only)}' — the site uses names like Red-Tiger, Pragmatic-Play." if only else ""
        return jsonify({"status": "error", "message":
                        "Nothing to test — every provider's games are in the 7-day rotation, "
                        "or no provider matched." + filt + (f" [{errs}]" if errs else "")}), 400

    started = datetime.now()
    ts = f"{started:%Y%m%d_%H%M%S}"
    batch_id = f"{ts}_DSCsweep"
    batch_dir = os.path.join(RUNS_DIR, batch_id)
    os.makedirs(batch_dir, exist_ok=True)

    worker_specs, group_meta = [], []
    total_games, widx = 0, 0
    for g in planned:
        b, r, picks, accts = g["brand"], g["region"], g["picks"], g["accounts"]
        gdir = os.path.join(batch_dir, f"{b}_{r}")
        os.makedirs(gdir, exist_ok=True)
        full_input = os.path.join(gdir, "input.xlsx")
        dsc_sweep._write_sheet(picks, full_input)
        report = os.path.join(RUNS_DIR, f"DSC_Sweep_{b}_{r}_{started:%Y-%m-%d_%H%M%S}.xlsx")
        dsc_report.ensure_report(report, seed_from=full_input)
        try:
            shards, n = dsc_report.shard_excel(full_input, gdir, len(accts))
        except Exception as e:
            _ = e; continue
        total_games += n
        for k, (shard, acc) in enumerate(zip(shards, accts), 1):
            widx += 1
            cmd = _worker_cmd("test_spin_button.py", "--excel", shard,
                   "--username", acc["username"], "--password", acc["password"],
                   "--brand", b, "--region", r,
                   "--run-dir", os.path.join(gdir, f"w{k}"),
                   "--tests", "dsc", "--dsc-report", report)
            if is_headless:
                cmd.append("--headless")
            env = os.environ.copy()
            step = 36 * ((widx - 1) % parallel)     # cascade within a wave; title bars stay clickable
            env["GAMEGUARD_WINDOW_POS"] = f"{step},{step}"
            worker_specs.append({"label": f"{r}·W{k}", "cmd": cmd, "env": env})
        group_meta.append({"brand": b, "region": r, "report": report, "picks": picks,
                           "record": not do_not_record, "run_id": batch_id})

    n_regions = len(group_meta)
    scope_label = (f"{n_regions} regions" if n_regions > 1
                   else f"{group_meta[0]['brand']} {group_meta[0]['region']}")
    CURRENT_RUN = {"game": f"Auto-sweep · {total_games} games · {scope_label}", "run_id": batch_id,
                   "start_time": started.isoformat(), "status": "running",
                   "batch": {"report": group_meta[0]["report"] if group_meta else None,
                             "total": total_games, "workers": len(worker_specs),
                             "run_id": batch_id}}
    try:
        index = []
        if os.path.exists(RUNS_INDEX):
            with open(RUNS_INDEX, "r", encoding="utf-8") as fh:
                index = json.load(fh)
        index.insert(0, {"run_id": batch_id, "game": f"Auto-sweep · {total_games} games · {scope_label}",
                         "brand": group_meta[0]["brand"], "region": group_meta[0]["region"],
                         "started_at": started.isoformat(), "workers": len(worker_specs),
                         "type": "sweep", "report": os.path.basename(group_meta[0]["report"])})
        with open(RUNS_INDEX, "w", encoding="utf-8") as fh:
            json.dump(index[:200], fh, indent=2)
    except Exception as e:
        print(f"[runs] index write failed: {e}")

    _start_fleet(worker_specs, parallel, group_meta)
    return jsonify({"status": "started", "total": total_games, "workers": len(worker_specs),
                    "parallel": parallel, "regions": n_regions, "run_id": batch_id,
                    "record": not do_not_record,
                    "groups": [{"brand": g["brand"], "region": g["region"],
                                "games": len(g["picks"])} for g in group_meta]})


@app.route('/api/crash-sweep-plan', methods=['POST'])
def api_crash_sweep_plan():
    """Preview every crash title available for a (brand, region) — ONE live-catalog call
    (provider_sweep.list_crash_games reads the Categories API's curated 'crashgames' bucket
    directly, unlike the slot sweep which must walk every provider). No spins, no money.
    Crash is explicitly OUT of the DSC weekly-rotation compliance pass slots use (see
    SLOT_TEST_CHECKLIST.md — 'handled outside the slot UI pass'), so this is a plain
    discover-and-list, not a rotation pick, and nothing here touches dsc_history_db.
    JSON body: {username, password, brand, region}."""
    data = request.get_json(force=True) or {}
    username, password = data.get('username', ''), data.get('password', '')
    brand = (data.get('brand') or 'betway').lower()
    region = (data.get('region') or 'ZA').upper()
    if not username or not password:
        return jsonify({"status": "error", "message": "Select an account first"}), 400
    try:
        token, err = _get_token(username, password, brand=brand, region=region)
        if not token:
            return jsonify({"status": "error", "message": f"auth failed: {err}"}), 400
        import modules.provider_sweep as provider_sweep
        games = provider_sweep.list_crash_games(brand, region, token)
    except Exception as e:
        return jsonify({"status": "error", "message": f"Discovery failed: {e}"}), 400
    return jsonify({"status": "ok", "games": games, "total": len(games)})


def _do_launch_crash_sweep(picks, username, password, brand, region, headless, live, bet, target):
    """Core of a crash sweep launch. Factored out of the /launch-crash-sweep route
    (2026-08-10) so the scheduler can trigger the EXACT same launch a manual click does —
    same validation (account required, MAX_STAKE cap), same one-at-a-time worker fleet.
    Dry-run by default; live=True places a REAL wager on EVERY selected game, one after
    another (2026-07-29, explicit request — "whatever I select in discover games should be
    considered in live execution"). Stake is re-validated against config_env.MAX_STAKE per
    game — the same cap the single-game Live card enforces, just applied N times here, and
    total exposure is surfaced in the response so it's never a silent multiply. Runs ONE AT A
    TIME regardless of dry-run or live — max_parallel is hard-capped to 1 on purpose: every
    game in the batch authenticates with the SAME account, and a real concurrent second login
    on one account is exactly the 'session held by another tab' disconnect crash_auto.py's
    auto_handle_crash_startup already has to detect and abort on (see
    modules/auth_handler.py's sessionTrackingToken note). That risk is WORSE, not better, in
    live mode — a dropped session mid-bet is a real open position, not just a wasted
    observation — so the one-at-a-time cap is non-negotiable here. Spreading a crash sweep
    across MULTIPLE accounts in parallel (like the slot Auto Sweep does) is future work — not
    built here, so don't silently pretend a 'parallel' knob exists.
    Stake is OPTIONAL: each game auto-wagers its OWN catalog minBetAmount (from
    `picks`, see /api/crash-sweep-plan -> list_crash_games) — `bet`, if given, overrides that
    for EVERY selected game instead (still capped by MAX_STAKE). `picks`:
    [{id,name,provider,minBetAmount}]. Returns (payload_dict, http_status)."""
    global CURRENT_RUN
    if not username or not password:
        return {"status": "error", "message": "Select an account first"}, 400
    picks = [p for p in (picks or []) if (p or {}).get('name')]
    if not picks:
        return {"status": "error", "message": "No crash games selected"}, 400

    is_live = bool(live)
    live_bet_override, live_target = 0, (target or '').strip()
    if is_live:
        import config_env
        try:
            live_bet_override = float(bet or 0)
        except (TypeError, ValueError):
            live_bet_override = 0
        if live_bet_override > 0 and live_bet_override > config_env.MAX_STAKE:
            return {"status": "error",
                    "message": f"Stake {live_bet_override:g} exceeds the safety cap "
                               f"{config_env.MAX_STAKE:g}"}, 400

    is_headless = bool(headless)
    started = datetime.now()
    batch_id = f"{started:%Y%m%d_%H%M%S}_crash_sweep"
    batch_dir = os.path.join(RUNS_DIR, batch_id)
    os.makedirs(batch_dir, exist_ok=True)

    worker_specs = []
    for i, p in enumerate(picks, 1):
        name = p["name"]
        game_min_bet = p.get("minBetAmount")
        gdir = os.path.join(batch_dir, f"{i:02d}_{_slug(name)}")
        os.makedirs(gdir, exist_ok=True)
        # Persist the discovery-time pick (id/name/provider/minBetAmount) alongside the run —
        # crash_auto.py's own results.json has no provider field, so the Excel exporter
        # (_crash_report_rows) needs this to fill that column without a second live catalog call.
        try:
            with open(os.path.join(gdir, "pick.json"), "w", encoding="utf-8") as fh:
                json.dump({"id": p.get("id"), "name": name, "provider": p.get("provider"),
                          "minBetAmount": game_min_bet}, fh)
        except Exception as e:
            print(f"[crash-sweep] pick.json write failed for {name}: {e}")
        cmd = _worker_cmd("crash_auto.py", "--game", name,
               "--username", username, "--password", password,
               "--brand", brand, "--region", region)
        if is_live:
            cmd += ["--live"]
            if live_bet_override > 0:
                cmd += ["--bet", str(live_bet_override)]
            elif game_min_bet not in (None, ""):
                cmd += ["--min-bet", str(game_min_bet)]
            if live_target:
                cmd += ["--target", live_target]
        else:
            # Dry-run only (run_crash_tests_with_retry refuses retries under --live — see its
            # docstring): retry up to 3x with a fresh launch URL on a session-drop or failed
            # control detection, since the 2026-08-03 provider validation sweep proved the SAME
            # game can flip between a clean pass and a false-negative failure across attempts.
            cmd += ["--dry-run", "--retries", "3"]
        cmd += ["--run-dir", gdir]
        if is_headless:
            cmd.append("--headless")
        worker_specs.append({"label": _slug(name, 16), "cmd": cmd, "env": None})

    total_games = len(worker_specs)
    total_exposure = None
    if is_live:
        if live_bet_override > 0:
            mode_label = f"LIVE (stake={live_bet_override:g} x {total_games} = " \
                        f"{live_bet_override*total_games:g} exposure)"
            total_exposure = live_bet_override * total_games
        else:
            known = [v for v in (_parse_money(p.get("minBetAmount")) for p in picks) if v is not None]
            if len(known) == total_games:
                total_exposure = sum(known)
                mode_label = f"LIVE (each game's own minimum bet = {total_exposure:g} exposure)"
            else:
                mode_label = "LIVE (each game's own minimum bet)"
    else:
        mode_label = "dry-run"
    CURRENT_RUN = {"game": f"Crash sweep · {total_games} games · {mode_label} · {brand} {region}",
                   "run_id": batch_id, "start_time": started.isoformat(), "status": "running",
                   "batch": {"report": None, "total": total_games, "workers": total_games,
                             "run_id": batch_id}}
    try:
        index = []
        if os.path.exists(RUNS_INDEX):
            with open(RUNS_INDEX, "r", encoding="utf-8") as fh:
                index = json.load(fh)
        index.insert(0, {"run_id": batch_id, "game": f"Crash sweep · {total_games} games · {mode_label}",
                         "brand": brand, "region": region,
                         "started_at": started.isoformat(), "workers": total_games,
                         "type": "crash-sweep", "report": None})
        with open(RUNS_INDEX, "w", encoding="utf-8") as fh:
            json.dump(index[:200], fh, indent=2)
    except Exception as e:
        print(f"[runs] index write failed: {e}")

    # max_parallel=1 (same-account risk, see docstring) + a cooldown gap between launches so the
    # casino backend's session lock for this account has time to release before the next game
    # requests a launch — see _start_fleet's cooldown docstring for the evidence behind this.
    _start_fleet(worker_specs, 1, [], cooldown=20)
    return {"status": "started", "total": total_games, "run_id": batch_id,
            "live": is_live, "total_exposure": total_exposure if is_live else 0}, 200


@app.route('/launch-crash-sweep', methods=['POST'])
def launch_crash_sweep():
    """Batch-run a set of discovered crash titles — see _do_launch_crash_sweep for the full
    behavior/safety docstring. JSON body: picks[{id,name,provider,minBetAmount}], username,
    password, brand, region, headless, live, bet, target."""
    data = request.get_json(force=True) or {}
    payload, status = _do_launch_crash_sweep(
        picks=data.get('picks'), username=data.get('username', ''), password=data.get('password', ''),
        brand=(data.get('brand') or 'betway').lower(), region=(data.get('region') or 'ZA').upper(),
        headless=bool(data.get('headless')), live=bool(data.get('live')),
        bet=data.get('bet'), target=data.get('target'))
    return jsonify(payload), status


@app.route('/api/history')
def api_history():
    """Prod-DB audit view for the History panel: most-recent tested games, optionally filtered."""
    from modules import dsc_history_db
    brand = request.args.get('brand') or None
    region = request.args.get('region') or None
    try:
        limit = min(int(request.args.get('limit', 100)), 500)
    except ValueError:
        limit = 100
    return jsonify({"rows": dsc_history_db.recent_rows(brand, region, limit)})


@app.route('/stream')
def stream():
    """Broadcast SSE: each client reads LOG_HISTORY from position 0, so a client that
    connects (or reconnects) mid-run still receives the whole log, and any number of
    tabs can watch the same run simultaneously."""
    def generate():
        while LOG_Q is None:
            time.sleep(0.3)
        pos = 0
        idle = 0.0
        while True:
            hist = list(LOG_HISTORY)
            if pos > len(hist):      # a new run cleared the log under us — hand off
                break
            if pos < len(hist):
                for line in hist[pos:]:
                    yield f"data: {line}\n\n"
                pos = len(hist)
                idle = 0.0
                continue
            if RUN_DONE["v"]:
                break
            time.sleep(0.25)
            idle += 0.25
            if idle >= 15:
                # SSE keep-alive comment so proxies/browsers don't drop an idle stream.
                yield ": ping\n\n"
                idle = 0.0
        yield "data: [GAMEGUARD_STREAM_END]\n\n"

    return Response(generate(), mimetype="text/event-stream",
                    headers={"Cache-Control": "no-cache", "X-Accel-Buffering": "no"})


@app.route('/stop', methods=['POST'])
def stop():
    global CURRENT_RUN
    FLEET_STOP.set()   # stop a capped fleet from launching further queued waves
    if _workers_alive():
        _stop_workers()
        CURRENT_RUN["status"] = "stopped"
        return jsonify({"status": "stopped"})
    return jsonify({"status": "idle"})


if __name__ == '__main__':
    # GAMEGUARD_RELOADER=0 (set by the QA "Launch GameGuard.bat") runs a single plain process —
    # no live-reload-on-save, but also no second watcher process for a non-developer to
    # accidentally leave orphaned after closing the window. Default (unset) keeps today's
    # developer behavior unchanged: debug=True's reloader re-execs this module in TWO
    # processes, a watcher (never serves — WERKZEUG_RUN_MAIN unset) and the actual serving
    # child (WERKZEUG_RUN_MAIN="true"). Gate the background thread so it starts EXACTLY once,
    # in whichever process actually serves: the single process when the reloader is off, else
    # only the flagged child — starting it in the watcher too would spawn a second, independent
    # scheduler racing to launch the same due jobs.
    use_reloader = os.environ.get("GAMEGUARD_RELOADER", "1") != "0"
    if not use_reloader or os.environ.get("WERKZEUG_RUN_MAIN") == "true":
        threading.Thread(target=_schedule_loop, daemon=True).start()
    app.run(debug=True, port=int(os.environ.get("GAMEGUARD_PORT", 5000)), threaded=True,
            use_reloader=use_reloader)
