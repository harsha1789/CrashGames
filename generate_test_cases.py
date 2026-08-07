"""
Generate a professional Excel test case document for the Slot Automation Framework.
Run: python generate_test_cases.py
Output: slot_automation_test_cases.xlsx
"""

import openpyxl
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side, GradientFill
)
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

wb = openpyxl.Workbook()

# ─── Color Palette ────────────────────────────────────────────────────────────
CLR_HEADER_BG   = "1E3A5F"   # dark navy
CLR_HEADER_FG   = "FFFFFF"
CLR_CAT_BG      = "2E75B6"   # medium blue
CLR_CAT_FG      = "FFFFFF"
CLR_ROW_ODD     = "EBF3FB"
CLR_ROW_EVEN    = "FFFFFF"
CLR_PASS        = "C6EFCE"
CLR_FAIL        = "FFC7CE"
CLR_SKIP        = "FFEB9C"
CLR_OPTIONAL    = "E2EFDA"
CLR_CRITICAL    = "FCE4D6"
CLR_SECTION_BG  = "D6E4F7"

def hfill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def bold(size=10, color="000000", italic=False):
    return Font(name="Calibri", size=size, bold=True, color=color, italic=italic)

def normal(size=10, color="000000"):
    return Font(name="Calibri", size=size, color=color)

def wrap_align(h="left", v="center"):
    return Alignment(horizontal=h, vertical=v, wrap_text=True)

thin = Side(style="thin", color="BDD7EE")
border = Border(left=thin, right=thin, top=thin, bottom=thin)

# ══════════════════════════════════════════════════════════════════════════════
#  SHEET 1 — Test Cases
# ══════════════════════════════════════════════════════════════════════════════
ws = wb.active
ws.title = "Test Cases"
ws.sheet_view.showGridLines = False

# ── Title banner ──────────────────────────────────────────────────────────────
ws.merge_cells("A1:J1")
title_cell = ws["A1"]
title_cell.value = "🎰  Slot Automation Framework — Test Case Specification"
title_cell.font  = Font(name="Calibri", size=16, bold=True, color=CLR_HEADER_FG)
title_cell.fill  = hfill(CLR_HEADER_BG)
title_cell.alignment = wrap_align("center")
ws.row_dimensions[1].height = 36

ws.merge_cells("A2:J2")
sub_cell = ws["A2"]
sub_cell.value = "Model: gemini-2.5-flash  |  Browser: Playwright (Chromium)  |  Viewport: 1920×1080  |  Total Core Tests: 17"
sub_cell.font  = Font(name="Calibri", size=10, italic=True, color="555555")
sub_cell.fill  = hfill("D9E8F5")
sub_cell.alignment = wrap_align("center")
ws.row_dimensions[2].height = 20

# ── Column headers ─────────────────────────────────────────────────────────────
headers = [
    "Test ID", "Category", "Test Name", "Description / Steps",
    "Expected Result", "Pass Criteria (Code)", "Priority",
    "Type", "Flags / Notes", "Status"
]
col_widths = [10, 22, 30, 60, 42, 38, 12, 16, 30, 12]

for col, (h, w) in enumerate(zip(headers, col_widths), 1):
    cell = ws.cell(row=3, column=col, value=h)
    cell.font      = bold(11, CLR_HEADER_FG)
    cell.fill      = hfill(CLR_CAT_BG)
    cell.alignment = wrap_align("center")
    cell.border    = border
    ws.column_dimensions[get_column_letter(col)].width = w

ws.row_dimensions[3].height = 30

# ── Test data ─────────────────────────────────────────────────────────────────
# Format: (test_id, category, name, description, expected, pass_criteria, priority, type_, flags)
tests = [

    # SETUP / PRE-FLIGHT
    ("TC-001", "Setup & Pre-flight",
     "Game Page Loads",
     "1. Navigate to slot game URL\n2. Wait for DOMContentLoaded\n3. Run AI startup loop (max 15 attempts)\n4. Detect 'ready' state via Gemini vision",
     "Game fully loads within 15 AI-check attempts. Reels, spin button, and bet amount are visible with no overlay blocking the screen.",
     "auto_handle_startup() returns True",
     "Critical", "Automated", "Uses Gemini to detect 3 states: wait / click / ready. Clicks any intro overlays automatically.", ""),

    ("TC-002", "Setup & Pre-flight",
     "Pre-flight Balance Check",
     "1. Take screenshot after game load\n2. Read balance and bet values using Gemini\n3. Compare balance vs bet amount",
     "Balance ≥ Bet. If balance < bet, test suite aborts with 'INSUFFICIENT FUNDS' message.",
     "pre_bal >= pre_bet (abort if not)",
     "Critical", "Automated", "Prevents false failures due to zero-balance accounts.", ""),

    # UI DETECTION
    ("TC-003", "UI Detection",
     "All UI Controls Detected",
     "1. Capture full-page screenshot\n2. Send to Gemini with structured detect prompt\n3. Parse JSON array of controls with box_2d coords",
     "At least 3 controls detected: Spin Button, Bet Display, Balance Display. Each has valid label and bounding box.",
     "len(controls) >= 3",
     "Critical", "Automated", "Gemini detects: Spin, Autoplay, Bet+, Bet-, Menu, Sound, Turbo, Info, BuyBonus, Balance, Bet displays.", ""),

    ("TC-004", "UI Detection",
     "Default Bet Verification",
     "1. Read bet amount from UI immediately after load\n2. Compare with expected default bet value passed via --default-bet flag",
     "UI-shown default bet matches the expected value configured for this game.",
     "pre_bet == parse_amount(default_bet)",
     "High", "Automated", "Optional test — only runs when --default-bet arg is provided.", "Optional"),

    ("TC-005", "UI Detection",
     "Minimum Bet Verification",
     "1. Spam-click Bet Decrement button 8 times\n2. Take screenshot\n3. Read resulting bet value\n4. Compare to expected minimum",
     "After 8 decrements, the displayed bet matches the configured minimum bet value.",
     "min_bet_found == parse_amount(min_bet)",
     "High", "Automated", "Optional — only runs when --min-bet arg is provided.", "Optional"),

    # NETWORK / SPIN ENDPOINT
    ("TC-006", "Network Interception",
     "Spin Endpoint Discovered",
     "1. Record all idle network traffic for 8s (baseline)\n2. Click spin button once\n3. Wait 8s for response\n4. Diff new requests against idle baseline\n5. First non-noise, non-idle POST/WS path = spin endpoint",
     "A unique spin API endpoint (HTTP POST or WebSocket path) is identified and stored.",
     "monitor.spin_endpoint is not None",
     "Critical", "Automated", "Handles both HTTP REST and WebSocket spin protocols. Noise filter excludes analytics, assets, fonts.", ""),

    ("TC-007", "Network Interception",
     "Server Returns Valid Spin Response",
     "1. Collect all HTTP responses matching spin endpoint\n2. Check last response status code\n3. Attempt JSON parse of response body",
     "Spin endpoint returns HTTP 200 with a valid JSON body on every spin.",
     "status == 200 and json.loads(body) succeeds",
     "Critical", "Automated", "Validates backend is responding correctly. Fails if response is non-JSON or non-200.", ""),

    # SPIN BUTTON
    ("TC-008", "Spin Button",
     "Single Click = 1 Spin Request",
     "1. Record baseline spin count = 0\n2. Click spin button once\n3. Wait for spin completion (network response received)\n4. Count spin requests intercepted",
     "Exactly 1 spin API request is triggered per single button click.",
     "monitor.spin_count == 1",
     "Critical", "Automated", "Core correctness test. Ensures no double-firing on single click.", ""),

    ("TC-009", "Spin Button",
     "Rapid Clicks During Spin = Still 1 Request",
     "1. Click spin button\n2. Once spin request is detected in-flight (within 3s)\n3. Spam-click spin button 8 more times\n4. Wait for completion\n5. Count total spin requests",
     "Despite 9 total clicks, only 1 spin request fires. Spin button becomes disabled while spin is in-progress.",
     "monitor.spin_count == 1 after 9 clicks",
     "Critical", "Automated", "Verifies the game prevents double-spend. Button must be disabled or debounced during active spin.", ""),

    ("TC-010", "Spin Button",
     "Spin Button Re-enables After Completion",
     "1. Click spin button\n2. Wait for full spin completion (response received + 1.5s settle)\n3. Click spin button again\n4. Count second spin requests",
     "A second spin fires successfully after the first completes. Spin count = 1 for the second click.",
     "monitor.spin_count == 1 (second spin)",
     "Critical", "Automated", "Verifies button re-activates after spin. Prevents game being stuck after one spin.", ""),

    # WAGER & BALANCE
    ("TC-011", "Wager & Balance",
     "Wager Correctly Processed",
     "1. Read balance and bet before spin\n2. Execute single spin\n3. Verify bet amount was readable before spin initiated",
     "A non-null bet amount is visible and readable before the spin. Confirms the wager is properly displayed.",
     "prespin_bet is not None",
     "High", "Automated", "Validates wager display is functional and parseable. Prerequisite for balance math test.", ""),

    ("TC-012", "Wager & Balance",
     "Balance Updated Correctly",
     "1. Read balance (B) and bet (W) before spin\n2. Execute spin\n3. Detect payout (P) from network response\n4. Read post-spin balance\n5. Verify: post_balance ≈ B - W + P",
     "Post-spin balance equals pre-spin balance minus wager plus any payout (±0.10 tolerance for floating point).",
     "abs(postspin_bal - (prespin_bal - prespin_bet + payout)) < 0.10",
     "Critical", "Automated", "Core financial integrity check. Tolerance of 0.10 accounts for floating-point display rounding.", ""),

    ("TC-013", "Wager & Balance",
     "Payout Successfully Logged",
     "1. Parse spin response JSON body\n2. Search for keys: win, winAmount, payout, totalWin\n3. Extract numeric payout value",
     "Payout amount (even if 0) is successfully extracted from network response. No errors parsing payout data.",
     "payout value extracted without exception",
     "Medium", "Automated", "Informational — always passes. Records payout for audit trail. Non-zero payout is logged separately.", ""),

    ("TC-014", "Wager & Balance",
     "Feature Triggers Monitored",
     "1. After spin, parse response body as string\n2. Check for keywords: feature, freespin, bonus\n3. Flag feature_triggered = True if found",
     "Feature trigger status is correctly detected from network response (True/False). No false positives.",
     "feature_triggered flag set correctly",
     "Medium", "Automated", "Informational — always passes. Detects bonus rounds, free spins, jackpot triggers for audit log.", ""),

    # BET CONTROLS
    ("TC-015", "Bet Controls",
     "Bet Can Be Changed",
     "Strategy A: Click Bet + button, read new bet, check if changed.\nStrategy B: Click Bet Display area → check for overlay → pick different bet value.\nStrategy C: Click Bet - button, read new bet, check if changed.\nFirst strategy that changes the bet value wins.",
     "Bet amount changes from its initial value using at least one of the three strategies.",
     "bet_after != bet_before (using any strategy)",
     "High", "Automated", "Adaptive multi-strategy approach handles both direct +/- buttons and bet overlay/popup patterns.", ""),

    ("TC-016", "Bet Controls",
     "Bet Can Be Restored (Round-trip)",
     "1. After changing bet (TC-015)\n2. Click restore button (Bet- or Bet+)\n3. If overlay appears, use Gemini to find original bet value option\n4. Click original bet value\n5. Verify bet returned to original",
     "Bet successfully returns to its original value after the change cycle (round-trip).",
     "abs(bet_restored - bet_before) < 0.01",
     "High", "Automated", "Only runs if TC-015 passed. Verifies bet change is reversible — no stuck state.", ""),

    ("TC-017", "Bet Controls",
     "Bet Buttons Disabled During Spin",
     "1. Read current bet value\n2. Start spin\n3. Once spin request is in-flight, click Bet +/- button twice\n4. Wait for spin to complete\n5. Read bet value after spin",
     "Bet amount does not change despite clicking Bet +/- during active spin. Bet buttons are disabled or ignored while spin is in progress.",
     "bet_before == bet_after (no change despite clicks during spin)",
     "Critical", "Automated", "Prevents mid-spin bet manipulation which could cause financial integrity issues.", ""),

    # AUTOPLAY
    ("TC-018", "Autoplay",
     "Auto Play Functionality Works",
     "1. Detect Autoplay button from control detection\n2. Click Autoplay button\n3. Wait 2s for response\n4. Click Autoplay again to disable\n5. Click screen center to dismiss any panel",
     "Autoplay button responds to click without crashing the game. Panel/menu may open or autoplay may start.",
     "No exception on click; game remains functional after interaction",
     "Medium", "Automated", "Skipped if Autoplay button not detected in TC-003. Interaction validated, not full autoplay cycle.", ""),

    # MENU / NAVIGATION
    ("TC-019", "Menu / Navigation",
     "Menu Button Opens a Panel",
     "1. Take screenshot BEFORE clicking menu\n2. Click Menu/hamburger button\n3. Wait 2s\n4. Take screenshot AFTER clicking menu\n5. Send both screenshots to Gemini\n6. Ask: 'Did a menu/settings panel open?'",
     "Gemini confirms a menu panel, settings overlay, or navigation panel visually appeared after the button click.",
     "menu_result['menu_opened'] == True",
     "High", "Automated", "Skipped if Menu button not detected. Uses 2-image Gemini comparison for visual diff verification.", ""),

    # REGION-WISE TESTS
    ("TC-020", "Region — UK",
     "GBP Currency Format Parsed",
     "Run full test suite with UK game URL.\nVerify balance and bet amounts display in GBP (£) format and are parsed correctly by parse_amount().",
     "Balance shows £X.XX format. parse_amount() returns correct float. No currency parsing errors.",
     "pre_bal is not None and pre_bet is not None",
     "High", "Automated", "UK regulatory requirement — GBP display. Verify responsible gambling overlay is handled.", "Region"),

    ("TC-021", "Region — South Africa",
     "ZAR Currency Format Parsed",
     "Run full test suite with ZA game URL.\nVerify 'R X.XX' or 'R X,XX' format parsed correctly.",
     "ZAR amounts (e.g. R 1 328.73) parsed to correct float. Decimal and thousands separators handled.",
     "pre_bal is not None, parse_amount('R 1 328.73') == 1328.73",
     "High", "Automated", "ZAR uses space as thousands separator. 'R' prefix must be stripped correctly.", "Region"),

    ("TC-022", "Region — Europe (DE/FR)",
     "European Decimal Format Parsed",
     "Run full test suite with European game URL.\nVerify amounts using comma as decimal separator (e.g. 1.000,50) are parsed correctly.",
     "European format (1.000,50 = 1000.50) parsed to correct float. No parse failures.",
     "parse_amount('1.000,50') == 1000.50",
     "High", "Automated", "European number format is opposite to UK/US. parse_amount() handles both conventions.", "Region"),

    ("TC-023", "Region — Canada",
     "CAD Currency + Responsible Gambling Banner",
     "Run full test suite with CA game URL.\nGame may show responsible gambling message on load — AI startup handler must click through it.",
     "Responsible gambling overlay is auto-dismissed by startup AI handler. Game loads fully. CAD amounts parsed correctly.",
     "auto_handle_startup() succeeds; pre_bal not None",
     "High", "Automated", "Canada requires RG banners. Startup AI loop must detect 'click' state and dismiss.", "Region"),

    ("TC-024", "Region — Australia",
     "AUD Currency + Strict RG Overlay",
     "Run full test suite with AU game URL.\nAustralia has strict responsible gambling requirements — typically 2-3 mandatory overlays before game starts.",
     "All mandatory overlays dismissed. Game loads to playable state. AUD amounts parsed correctly.",
     "auto_handle_startup() returns True within 15 attempts",
     "High", "Automated", "AU may have more overlays than other regions. Max startup retries may need to increase.", "Region"),

    ("TC-025", "Region — India",
     "Mobile Viewport + INR Currency",
     "Run full test suite with --mobile flag (iPhone 13 viewport).\nVerify game works on mobile viewport. INR (₹) currency parsed correctly.",
     "Game loads correctly in 390×844 mobile viewport. ₹ currency parsed correctly. Spin, bet controls visible and functional.",
     "All core tests pass in mobile mode; pre_bal parsed from ₹ format",
     "High", "Automated", "Uses Playwright iPhone 13 device preset. Mobile layout may differ significantly from desktop.", "Region"),

    ("TC-026", "Region — Brazil",
     "BRL Currency + Portuguese UI",
     "Run full test suite with BR game URL.\nPortuguese language overlays auto-handled by startup AI. BRL (R$) amounts parsed.",
     "Portuguese-language overlays (e.g. 'Continuar', 'Jogar') detected and clicked by AI. BRL amounts parsed correctly.",
     "auto_handle_startup() succeeds; R$ amounts parsed correctly",
     "Medium", "Automated", "Portuguese UI text — AI startup handles language-agnostic button detection via visual analysis.", "Region"),

    # MOBILE
    ("TC-027", "Mobile",
     "Mobile Layout — Controls Detected",
     "Run test suite with --mobile flag.\n1. Playwright uses iPhone 13 device profile (390×844)\n2. AI detects controls in mobile layout\n3. All spin/bet tests run as normal",
     "Minimum 3 UI controls detected in mobile layout. Spin button functional. Network spin endpoint discovered.",
     "len(controls) >= 3 in mobile viewport",
     "High", "Automated", "Mobile layout may have different button positions, sizes, and arrangement than desktop.", "Mobile"),

    ("TC-028", "Mobile",
     "Mobile Spin Button Tap Works",
     "1. Launch in mobile viewport\n2. Detect spin button coordinates (scaled to 390×844)\n3. Tap spin button\n4. Verify spin endpoint triggered",
     "Single tap triggers exactly 1 spin request. Coordinates correctly mapped from AI detection to mobile viewport.",
     "monitor.spin_count == 1 in mobile mode",
     "High", "Automated", "Coordinate scaling must account for mobile viewport dimensions vs desktop detection.", "Mobile"),

    # EDGE CASES
    ("TC-029", "Edge Cases",
     "API Key Rotation on Rate Limit",
     "1. Trigger Gemini 429/quota error by exhausting one key\n2. Verify rotate_api_key() switches to next key\n3. Retry succeeds",
     "On API rate limit (429/QUOTA_EXCEEDED), system automatically rotates to next key and retries within 2–10s.",
     "gemini_call() succeeds after rotate_api_key() triggered",
     "High", "Automated", "8 API keys pooled. Rotation prevents test suite from failing due to single key exhaustion.", ""),

    ("TC-030", "Edge Cases",
     "Game Load Timeout Handling",
     "1. Load a game URL\n2. After 15 AI startup checks (45s), game still shows loading\n3. Verify suite proceeds rather than hanging",
     "If game never reaches 'ready' state, suite proceeds after max attempts with a warning. Does not hang indefinitely.",
     "auto_handle_startup() returns False after max_attempts; suite continues",
     "Medium", "Automated", "Prevents infinite hang on broken game URLs.", ""),

    ("TC-031", "Edge Cases",
     "Insufficient Funds Abort",
     "1. Test with account where balance < minimum bet\n2. Pre-flight check reads balance and bet\n3. Verify suite aborts gracefully",
     "Suite detects balance < bet, prints INSUFFICIENT FUNDS, closes browser cleanly. No spin attempted.",
     "pre_bal < pre_bet → return results early",
     "Critical", "Automated", "Prevents real-money or test-account overspend during automation.", ""),

    ("TC-032", "Edge Cases",
     "WebSocket Spin Protocol Detection",
     "1. Load a game using WebSocket for spin communication\n2. Trigger spin\n3. Verify WS frame captured as spin endpoint",
     "WebSocket spin frames are captured and treated as spin endpoint. Spin count correctly increments on WS send.",
     "monitor.spin_endpoint ends with @WS_SEND",
     "High", "Automated", "Handles both REST (HTTP POST) and WebSocket game providers transparently.", ""),

    ("TC-033", "Edge Cases",
     "Excel Bulk Test Queue Processing",
     "1. Provide Excel file with game names via --excel flag\n2. Script authenticates with casino backend\n3. Resolves each game to iframe URL\n4. Runs full test suite per game",
     "All games in Excel queue are processed in order. Failed game URL lookups are skipped with error logged. Results saved per game.",
     "Each game in parse_excel() queue attempted; suite runs for resolved iframe URLs",
     "High", "Automated", "Bulk testing mode — processes N games from spreadsheet automatically.", ""),
]

# ── Write test rows ─────────────────────────────────────────────────────────────
for row_num, test in enumerate(tests, start=4):
    tid, cat, name, desc, expected, criteria, priority, ttype, flags, *_ = test
    values = [tid, cat, name, desc, expected, criteria, priority, ttype, flags, ""]

    is_odd  = (row_num % 2 == 0)
    row_bg  = CLR_ROW_ODD if is_odd else CLR_ROW_EVEN

    # Priority color override
    if priority == "Critical":
        row_bg = CLR_CRITICAL
    elif flags == "Optional":
        row_bg = CLR_OPTIONAL
    elif flags == "Region":
        row_bg = "EAF4FF"
    elif flags == "Mobile":
        row_bg = "F0EBF8"

    for col, val in enumerate(values, 1):
        cell = ws.cell(row=row_num, column=col, value=val)
        cell.fill      = hfill(row_bg)
        cell.border    = border
        cell.alignment = wrap_align("left" if col not in [1, 7, 8, 10] else "center")
        cell.font      = normal(9)

    ws.row_dimensions[row_num].height = 72

    # Style Test ID column
    ws.cell(row=row_num, column=1).font = bold(9, "1E3A5F")
    # Style category
    ws.cell(row=row_num, column=2).font = bold(9, "2E75B6")
    # Style test name
    ws.cell(row=row_num, column=3).font = bold(9)

# ── Freeze panes ───────────────────────────────────────────────────────────────
ws.freeze_panes = "A4"

# ── Auto-filter ───────────────────────────────────────────────────────────────
ws.auto_filter.ref = f"A3:J{3 + len(tests)}"

# ══════════════════════════════════════════════════════════════════════════════
#  SHEET 2 — Summary Dashboard
# ══════════════════════════════════════════════════════════════════════════════
ws2 = wb.create_sheet("Summary")
ws2.sheet_view.showGridLines = False

ws2.merge_cells("A1:F1")
c = ws2["A1"]
c.value = "Slot Automation — Test Coverage Summary"
c.font  = Font(name="Calibri", size=14, bold=True, color=CLR_HEADER_FG)
c.fill  = hfill(CLR_HEADER_BG)
c.alignment = wrap_align("center")
ws2.row_dimensions[1].height = 32

# Category breakdown
categories = {}
for t in tests:
    cat = t[1].split("—")[0].strip()
    categories[cat] = categories.get(cat, 0) + 1

priority_count = {}
for t in tests:
    p = t[6]
    priority_count[p] = priority_count.get(p, 0) + 1

ws2["A3"] = "Category Breakdown"
ws2["A3"].font = bold(11, CLR_HEADER_FG)
ws2["A3"].fill = hfill(CLR_CAT_BG)
ws2["B3"] = "Count"
ws2["B3"].font = bold(11, CLR_HEADER_FG)
ws2["B3"].fill = hfill(CLR_CAT_BG)

r = 4
for cat, cnt in sorted(categories.items()):
    ws2.cell(row=r, column=1, value=cat).font   = normal(10)
    ws2.cell(row=r, column=1).fill              = hfill(CLR_ROW_ODD if r%2==0 else CLR_ROW_EVEN)
    ws2.cell(row=r, column=2, value=cnt).font   = bold(10)
    ws2.cell(row=r, column=2).fill              = hfill(CLR_ROW_ODD if r%2==0 else CLR_ROW_EVEN)
    ws2.cell(row=r, column=2).alignment         = wrap_align("center")
    r += 1

r += 1
ws2.cell(row=r, column=1, value="TOTAL").font  = bold(11)
ws2.cell(row=r, column=2, value=len(tests)).font = bold(11, "1E3A5F")
ws2.cell(row=r, column=2).alignment = wrap_align("center")

r += 2
ws2.cell(row=r, column=1, value="Priority Breakdown").font = bold(11, CLR_HEADER_FG)
ws2.cell(row=r, column=1).fill = hfill(CLR_CAT_BG)
ws2.cell(row=r, column=2, value="Count").font = bold(11, CLR_HEADER_FG)
ws2.cell(row=r, column=2).fill = hfill(CLR_CAT_BG)
r += 1
pcolors = {"Critical": CLR_CRITICAL, "High": "FFF2CC", "Medium": CLR_OPTIONAL}
for priority, cnt in sorted(priority_count.items()):
    ws2.cell(row=r, column=1, value=priority).font  = bold(10)
    ws2.cell(row=r, column=1).fill = hfill(pcolors.get(priority, CLR_ROW_EVEN))
    ws2.cell(row=r, column=2, value=cnt).font = bold(10)
    ws2.cell(row=r, column=2).fill = hfill(pcolors.get(priority, CLR_ROW_EVEN))
    ws2.cell(row=r, column=2).alignment = wrap_align("center")
    r += 1

# Quick-ref on column D
ws2["D3"] = "Gemini API Calls per Game"
ws2["D3"].font = bold(11, CLR_HEADER_FG)
ws2["D3"].fill = hfill(CLR_CAT_BG)
ws2.merge_cells("D3:F3")

api_calls = [
    ("Startup check loop", "×3 avg", "CLR_ROW_ODD"),
    ("Detect all controls", "×1", "CLR_ROW_EVEN"),
    ("Pre-flight read values", "×1", "CLR_ROW_ODD"),
    ("Pre-spin read values", "×1", "CLR_ROW_EVEN"),
    ("Post-spin read values", "×1", "CLR_ROW_ODD"),
    ("Bet before read", "×1", "CLR_ROW_EVEN"),
    ("Overlay comparison (Strategies A/B/C)", "×0–3", "CLR_ROW_ODD"),
    ("Bet after read", "×1", "CLR_ROW_EVEN"),
    ("Bet during spin read ×2", "×2", "CLR_ROW_ODD"),
    ("Menu before/after compare", "×1", "CLR_ROW_EVEN"),
    ("TOTAL (avg per game)", "~13–15 calls", "CLR_CAT_BG"),
]
for i, (desc, cnt, bg) in enumerate(api_calls, start=4):
    bg_hex = {"CLR_ROW_ODD": CLR_ROW_ODD, "CLR_ROW_EVEN": CLR_ROW_EVEN, "CLR_CAT_BG": CLR_CAT_BG}[bg]
    ws2.cell(row=i, column=4, value=desc).font  = (normal if bg != "CLR_CAT_BG" else bold)(10)
    ws2.cell(row=i, column=4).fill = hfill(bg_hex)
    ws2.cell(row=i, column=5, value=cnt).font   = bold(10, "1E3A5F" if i < 14 else CLR_HEADER_FG)
    ws2.cell(row=i, column=5).fill = hfill(bg_hex)
    ws2.cell(row=i, column=5).alignment = wrap_align("center")

for col, w in zip("ABCDEF", [38, 10, 4, 48, 12, 4]):
    ws2.column_dimensions[col].width = w

# ══════════════════════════════════════════════════════════════════════════════
#  SHEET 3 — Region Matrix
# ══════════════════════════════════════════════════════════════════════════════
ws3 = wb.create_sheet("Region Matrix")
ws3.sheet_view.showGridLines = False

ws3.merge_cells("A1:I1")
c = ws3["A1"]
c.value = "Region-wise Test Coverage Matrix"
c.font  = Font(name="Calibri", size=13, bold=True, color=CLR_HEADER_FG)
c.fill  = hfill(CLR_HEADER_BG)
c.alignment = wrap_align("center")
ws3.row_dimensions[1].height = 28

regions = ["🇬🇧 UK", "🇿🇦 South Africa", "🇨🇦 Canada", "🇩🇪 Germany", "🇮🇳 India", "🇦🇺 Australia", "🇳🇿 NZ", "🇧🇷 Brazil"]
region_tests = [
    ("Core Spin Test",         ["✅","✅","✅","✅","✅","✅","✅","✅"]),
    ("Balance Update",         ["✅","✅","✅","✅","✅","✅","✅","✅"]),
    ("Bet Change",             ["✅","✅","✅","✅","✅","✅","✅","✅"]),
    ("Menu Test",              ["✅","✅","✅","✅","✅","✅","✅","✅"]),
    ("Currency Parsing",       ["GBP £","ZAR R","CAD $","EUR €","INR ₹","AUD $","NZD $","BRL R$"]),
    ("RG Overlay Handling",    ["✅","⚠️","✅","✅","⚠️","✅ Strict","⚠️","✅"]),
    ("Mobile Viewport",        ["✅","✅","⚠️","⚠️","✅","✅","⚠️","⚠️"]),
    ("Default Bet Check",      ["✅","✅","✅","✅","✅","✅","✅","✅"]),
    ("Min Bet Check",          ["✅","✅","✅","✅","✅","✅","✅","✅"]),
    ("Estimated Games",        ["800","600","500","400","400","400","300","300"]),
    ("Est. Cost (Gemini Flash)",["$4.00","$3.30","$2.50","$2.40","$2.20","$2.60","$1.65","$1.80"]),
]

# Header row
ws3.cell(row=2, column=1, value="Test Area").font = bold(10, CLR_HEADER_FG)
ws3.cell(row=2, column=1).fill = hfill(CLR_CAT_BG)
ws3.column_dimensions["A"].width = 30

for col, reg in enumerate(regions, 2):
    cell = ws3.cell(row=2, column=col, value=reg)
    cell.font = bold(9, CLR_HEADER_FG)
    cell.fill = hfill(CLR_CAT_BG)
    cell.alignment = wrap_align("center")
    ws3.column_dimensions[get_column_letter(col)].width = 18

for ri, (area, vals) in enumerate(region_tests, start=3):
    is_odd = ri % 2 == 0
    bg = CLR_ROW_ODD if is_odd else CLR_ROW_EVEN
    ws3.cell(row=ri, column=1, value=area).font = bold(9)
    ws3.cell(row=ri, column=1).fill = hfill(bg)
    ws3.row_dimensions[ri].height = 22
    for col, v in enumerate(vals, 2):
        cell = ws3.cell(row=ri, column=col, value=v)
        cell.alignment = wrap_align("center")
        cell.font = normal(9)
        if v == "✅":
            cell.fill = hfill(CLR_PASS)
        elif "⚠️" in str(v):
            cell.fill = hfill(CLR_SKIP)
        elif str(v).startswith("$"):
            cell.fill = hfill("E2EFDA")
            cell.font = bold(9, "375623")
        else:
            cell.fill = hfill(bg)

# ─── Save ──────────────────────────────────────────────────────────────────────
out_path = r"d:\demo-slot-ai\slot-auto\slot_automation_test_cases.xlsx"
wb.save(out_path)
print(f"\n[OK] Excel file saved: {out_path}")
print(f"   Sheets: 'Test Cases' ({len(tests)} tests), 'Summary', 'Region Matrix'")
