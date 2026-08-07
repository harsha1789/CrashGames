# modules/dsc_report.py
"""DSC daily report — SAME file format in and out (Sr. No. | Provider | Game Name |
Launch | Bet Placed | Tlogs | Error | Evidence).

For batch sweeps the report is seeded as a copy of the INPUT sheet (all rows, original
order, result columns cleared), and each game's row is filled in as its run completes —
so a partially-finished sweep is a valid, readable report at any moment. Single-game
runs upsert into the shared daily file (created with just the header if missing):
a re-run of the same game UPDATES its row rather than appending a duplicate.

Screenshots/video stay internal in the per-game run folder; the Evidence column carries
that folder's name so a row can be traced back to its artifacts (blank the column before
sending the sheet out if needed).
"""
import os
import json
import time
import shutil
from datetime import date
from contextlib import contextmanager

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font

DSC_COLUMNS = ["Sr. No.", "Provider", "Game Name", "Launch",
               "Bet Placed", "Tlogs", "Error", "Evidence"]
RESULT_COLUMNS = ["Launch", "Bet Placed", "Tlogs", "Error", "Evidence"]
_COL_WIDTHS = {"Sr. No.": 8, "Provider": 22, "Game Name": 36, "Launch": 10,
               "Bet Placed": 12, "Tlogs": 10, "Error": 46, "Evidence": 34}


def _norm(name) -> str:
    """Normalize a header for matching: 'Sr. No.' == 'Sr No' == 'srno'."""
    return "".join(ch for ch in str(name or "").lower() if ch.isalnum())


@contextmanager
def _locked(path, timeout=30.0, stale=120.0):
    """Cross-process lock via an exclusive lockfile. N parallel DSC workers share one
    report file, and openpyxl's load→modify→save is not atomic — an unlocked concurrent
    save corrupts the workbook. A lock older than `stale` seconds is treated as left
    behind by a killed worker and broken."""
    lock = str(path) + ".lock"
    deadline = time.time() + timeout
    while True:
        try:
            fd = os.open(lock, os.O_CREAT | os.O_EXCL | os.O_WRONLY)
            os.write(fd, str(os.getpid()).encode())
            os.close(fd)
            break
        except FileExistsError:
            try:
                if time.time() - os.path.getmtime(lock) > stale:
                    os.remove(lock)
                    continue
            except OSError:
                pass
            if time.time() > deadline:
                raise TimeoutError(f"report lock busy for {timeout}s: {lock}")
            time.sleep(0.08)
    try:
        yield
    finally:
        try:
            os.remove(lock)
        except OSError:
            pass


# Header aliases: team sheets don't always use the canonical spelling.
_ALIASES = {
    "Game Name": ("gamename", "game", "name"),
    "Sr. No.": ("srno", "sno", "serial"),
    "Provider": ("provider", "providername", "studio"),
}


def _aliases(canonical):
    return _ALIASES.get(canonical, (_norm(canonical),))


def _row_map(ws, r) -> dict:
    """{normalized header -> 1-based column index} from row r."""
    return {_norm(c.value): c.column for c in ws[r] if c.value is not None}


def _col(cols: dict, canonical: str):
    """1-based column index for a canonical DSC column, honoring aliases."""
    return next((cols[k] for k in _aliases(canonical) if k in cols), None)


def _find_header(ws, max_scan=15):
    """(header_row_idx, header_map). Team sheets put a title and/or blank rows ABOVE
    the header — pandas and the dashboard's SheetJS skip those silently, but openpyxl
    reads literal row 1, so scan the top rows for one containing a Game Name variant."""
    for r in range(1, min(max_scan, ws.max_row or 1) + 1):
        m = _row_map(ws, r)
        if _col(m, "Game Name"):
            return r, m
    return 1, _row_map(ws, 1)


def _pick_sheet(wb):
    """(worksheet, header_row, header_map) — the first sheet that actually has a Game
    Name column. wb.active is untrustworthy: it's whichever tab was open when the file
    was last saved (a notes tab, often)."""
    for cand in wb.worksheets:
        hdr, m = _find_header(cand)
        if _col(m, "Game Name"):
            return cand, hdr, m
    ws = wb.active
    hdr, m = _find_header(ws)
    return ws, hdr, m


def _ensure_columns(ws, hdr_row=1):
    """Make sure every DSC column exists on the header row (missing ones are appended
    at the right end; a column present under an alias is NOT duplicated)."""
    have = _row_map(ws, hdr_row)
    for col in DSC_COLUMNS:
        if not _col(have, col):
            idx = ws.max_column + 1
            ws.cell(row=hdr_row, column=idx, value=col).font = Font(bold=True)
            ws.column_dimensions[ws.cell(row=hdr_row, column=idx).column_letter].width = \
                _COL_WIDTHS.get(col, 14)
            have[_norm(col)] = idx
    return _row_map(ws, hdr_row)


def default_report_path(base_dir: str) -> str:
    """One shared file per day, so chunked runs of the daily 300 land in the same sheet."""
    return os.path.join(base_dir, "runs", f"DSC_Report_{date.today():%Y-%m-%d}.xlsx")


def ensure_report(path: str, seed_from: str = None) -> str:
    """Create the report if missing: as a copy of the input sheet (result columns
    cleared — stale Pass/Fail values from a previous day must not read as today's)
    when `seed_from` is given, else with just the header row."""
    with _locked(path):
        return _ensure_report_unlocked(path, seed_from)


def _ensure_report_unlocked(path: str, seed_from: str = None) -> str:
    if os.path.exists(path):
        return path
    os.makedirs(os.path.dirname(os.path.abspath(path)), exist_ok=True)
    if seed_from and os.path.exists(seed_from):
        shutil.copyfile(seed_from, path)
        wb = load_workbook(path)
        ws, hdr, _ = _pick_sheet(wb)
        cols = _ensure_columns(ws, hdr)
        for name in RESULT_COLUMNS:
            c = _col(cols, name)
            for r in range(hdr + 1, ws.max_row + 1):
                # NB: ws.cell(..., value=None) silently ignores None; assign to clear.
                ws.cell(row=r, column=c).value = None
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = "DSC"
        ws.append(DSC_COLUMNS)
        for cell in ws[1]:
            cell.font = Font(bold=True)
            ws.column_dimensions[cell.column_letter].width = _COL_WIDTHS.get(cell.value, 14)
    wb.save(path)
    return path


def upsert_row(path: str, row: dict) -> str:
    """Fill the game's row in place (matched by Game Name, disambiguated by Sr. No. when
    both sides have one); append a full row when the game isn't in the sheet yet."""
    with _locked(path):
        _ensure_report_unlocked(path)
        wb = load_workbook(path)
        ws, hdr, _ = _pick_sheet(wb)
        cols = _ensure_columns(ws, hdr)
        name_c, sr_c = _col(cols, "Game Name"), _col(cols, "Sr. No.")

        want_name = str(row.get("Game Name", "")).strip().lower()
        want_sr = str(row.get("Sr. No.", "")).strip()
        target = None
        for r in range(hdr + 1, ws.max_row + 1):
            if str(ws.cell(row=r, column=name_c).value or "").strip().lower() != want_name:
                continue
            have_sr = str(ws.cell(row=r, column=sr_c).value or "").strip()
            if want_sr and have_sr and have_sr != want_sr:
                continue   # same name listed twice — keep looking for the right serial
            target = r
            break

        if target is None:
            # Append positionally against the sheet's ACTUAL layout (its column order
            # may differ from DSC_COLUMNS).
            r = ws.max_row + 1
            for c in DSC_COLUMNS:
                idx = _col(cols, c)
                if idx:
                    ws.cell(row=r, column=idx, value=row.get(c, ""))
        else:
            for name in RESULT_COLUMNS:
                ws.cell(row=target, column=_col(cols, name), value=row.get(name, ""))
            # Fill Provider only if the sheet doesn't already say (the input sheet wins).
            prov_c = _col(cols, "Provider")
            if not str(ws.cell(row=target, column=prov_c).value or "").strip():
                ws.cell(row=target, column=prov_c, value=row.get("Provider", ""))
        wb.save(path)
        return path


# Backward-compatible name: earlier call sites appended; upsert supersedes it.
append_row = upsert_row


def update_result(path: str, game_name: str, sr_no, updates: dict,
                  append_error: str = None) -> bool:
    """Partially update a game's row (only the given columns) — used by the tlogs
    validator to fill the Tlogs verdict without touching Launch/Bet Placed. Returns
    False when the game isn't in the sheet. `append_error` merges into the Error
    column instead of overwriting the sweep's own notes."""
    with _locked(path):
        wb = load_workbook(path)
        ws, hdr, _ = _pick_sheet(wb)
        cols = _ensure_columns(ws, hdr)
        name_c, sr_c = _col(cols, "Game Name"), _col(cols, "Sr. No.")

        want_name = str(game_name or "").strip().lower()
        want_sr = str(sr_no if sr_no is not None else "").strip()
        target = None
        for r in range(hdr + 1, ws.max_row + 1):
            if str(ws.cell(row=r, column=name_c).value or "").strip().lower() != want_name:
                continue
            have_sr = str(ws.cell(row=r, column=sr_c).value or "").strip()
            if want_sr and have_sr and have_sr != want_sr:
                continue
            target = r
            break
        if target is None:
            return False

        for col_name, value in updates.items():
            c = _col(cols, col_name)
            if c:
                ws.cell(row=target, column=c).value = value
        if append_error:
            e = _col(cols, "Error")
            old = str(ws.cell(row=target, column=e).value or "").strip()
            # Idempotent: a re-run of the validator must not stack the same note again.
            if append_error not in old:
                merged = append_error if old in ("", "NA") else f"{old}; {append_error}"
                ws.cell(row=target, column=e).value = merged
        wb.save(path)
        return True


def records_path(report_path: str) -> str:
    """The JSONL bet-record file that pairs with a report. Every DSC spin appends one
    line here — the raw material for the later transaction-history verification pass
    (Betway back office reflects bets ~10-15 min late, so verification is a second
    script run over these records, not part of the live sweep)."""
    return os.path.splitext(report_path)[0] + "_records.jsonl"


def append_record(report_path: str, record: dict) -> str:
    """Append one bet record to the report's JSONL (lock shared across workers)."""
    p = records_path(report_path)
    with _locked(p):
        with open(p, "a", encoding="utf-8") as f:
            f.write(json.dumps(record, ensure_ascii=False, default=str) + "\n")
    return p


def shard_excel(input_path: str, out_dir: str, n: int):
    """Split the input sheet into up to `n` round-robin shard files (header preserved
    verbatim) so parallel workers each sweep a slice. Round-robin — not contiguous
    chunks — so slow providers don't all land on one worker. Rows without a Game Name
    are dropped (team sheets have trailing blanks). Returns (shard_paths, total_games)."""
    wb = load_workbook(input_path, data_only=True)
    ws, hdr_row, header_map = _pick_sheet(wb)
    header = [c.value for c in ws[hdr_row]]
    name_idx = _col(header_map, "Game Name")
    if not name_idx:
        # Self-diagnosing: show what headers each tab actually has, so the next
        # "column not found" isn't a mystery.
        peek = "; ".join(f"'{c.title}': {[str(v.value) for v in c[1] if v.value is not None][:8]}"
                         for c in wb.worksheets)
        raise ValueError(f'no "Game Name" column found in any sheet — headers seen: {peek}')
    name_idx -= 1
    rows = [list(r) for r in ws.iter_rows(min_row=hdr_row + 1, values_only=True)
            if len(r) > name_idx and r[name_idx] is not None and str(r[name_idx]).strip()]
    if not rows:
        raise ValueError("the input sheet has no game rows")
    n = max(1, min(int(n), len(rows)))
    os.makedirs(out_dir, exist_ok=True)
    paths = []
    for k in range(n):
        chunk = rows[k::n]
        if not chunk:
            continue
        swb = Workbook()
        sws = swb.active
        sws.title = "games"
        sws.append(header)
        for r in chunk:
            sws.append(r)
        p = os.path.join(out_dir, f"shard_{k + 1}.xlsx")
        swb.save(p)
        paths.append(p)
    return paths, len(rows)


def failure_row(meta: dict, reason: str) -> dict:
    """Row for a game that never reached the browser (not in catalog / iframe failed)."""
    return {
        "Sr. No.": meta.get("srNo", ""),
        "Provider": meta.get("provider") or "Unknown",
        "Game Name": meta.get("gameName", ""),
        "Launch": "Fail",
        "Bet Placed": "NA",
        "Tlogs": "NA",
        "Error": reason,
        "Evidence": meta.get("evidence", ""),
    }
